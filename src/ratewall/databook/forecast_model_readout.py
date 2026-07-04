"""Ten-year TDCSim/CBO forecast readout with timed beta paths."""

from __future__ import annotations

import csv
import json
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.request import urlretrieve

from ratewall.accounting.assumption_engine import load_ratewall_assumption_sets
from ratewall.databook.model_artifact_store import (
    ArtifactManifestView,
    artifact_manifest_exists,
)
from ratewall.databook.tdcsim_cbo_contracts import (
    BETA_CHI_ROBUSTNESS_BETA_PROFILES,
    DEFAULT_TDC_BETA,
    DEFAULT_TDC_DEPOSIT_CURRENT_DEMAND_SHARE,
    tdcsim_cbo_model_scenario_interpretation_synthesis_rows,
    tdcsim_cbo_model_scenario_materiality_classification_rows,
    tdcsim_cbo_model_scenario_summary_rows,
)

DEFAULT_FORECAST_READOUT_SUITE_DIR = Path(
    "var/tdcsim_cbo_suite_20260627_tdcsim72dc6c7_full10y_core"
)

TIMED_BETA_PATH_FIELDS = [
    "forecast_timed_beta_path_row_id",
    "fiscal_year",
    "scenario_id",
    "baseline_scenario_id",
    "beta_path_id",
    "beta_path_label",
    "beta_transition_fiscal_year",
    "tdc_materialization_beta_scenario",
    "tdc_materialization_beta",
    "tdc_materialization_beta_source_status",
    "deposit_current_demand_share",
    "derived_beta_times_chi",
    "tdc_change_ex_overlap_bil",
    "baseline_tdc_change_ex_overlap_bil",
    "tdc_current_demand_support_bil_recomputed",
    "delta_tdc_current_demand_support_bil_recomputed",
    "direct_treasury_current_demand_support_bil_fixed",
    "delta_direct_treasury_current_demand_support_bil_fixed",
    "bank_treasury_current_demand_support_bil_fixed",
    "delta_bank_treasury_current_demand_support_bil_fixed",
    "total_current_demand_support_bil_recomputed",
    "delta_total_current_demand_support_bil_recomputed",
    "selected_moving_denominator_bil",
    "selected_delta_denominator_bil",
    "level_ratewall_ratio_recomputed",
    "delta_ratewall_ratio_vs_baseline_recomputed",
    "delta_ratewall_ratio_vs_normal_forward_path",
    "wall_hit_under_timed_beta_path",
    "scenario_axis",
    "model_relevance_class",
    "allowed_use",
    "blocked_use",
    "claim_boundary",
    "canonical_ratio_entry",
    "enters_main_ratio",
    "evidence_mode_enabled",
]

FORECAST_CHANNEL_CLASSIFICATION_FIELDS = [
    "forecast_channel_classification_row_id",
    "channel_id",
    "channel_label",
    "current_or_historical_channel_status",
    "first_10y_forecast_status",
    "first_forecast_entry_role",
    "selected_central_entry_role",
    "public_interest_block_role",
    "classification",
    "reason",
    "next_model_requirement",
    "allowed_use",
    "blocked_use",
]

FORECAST_NUMERATOR_CHANNEL_PLAN_FIELDS = [
    "forecast_numerator_channel_plan_row_id",
    "channel_id",
    "channel_label",
    "current_classification",
    "final_central_status",
    "materiality_tier",
    "forecast_route",
    "assumption_basis",
    "calibration_need",
    "double_count_guard",
    "admission_test",
    "next_model_action",
    "allowed_use",
    "blocked_use",
]

ZERO_LOW_APR_CREDIT_MATERIALITY_FIELDS = [
    "zero_low_apr_credit_materiality_row_id",
    "product_segment",
    "source_vintage",
    "source_url",
    "source_metric",
    "annual_originations_bil",
    "outstanding_stock_bil",
    "zero_low_apr_share",
    "duration_months",
    "candidate_rate_wedge_pct",
    "screen_relief_bil",
    "screen_status",
    "central_n_treatment",
    "materiality_decision",
    "next_model_action",
    "allowed_use",
    "blocked_use",
]

PUBLIC_INTEREST_NET_BLOCK_FIELDS = [
    "forecast_public_interest_net_block_row_id",
    "fiscal_year",
    "scenario_id",
    "baseline_scenario_id",
    "assumption_set",
    "source_vintage",
    "direct_treasury_current_demand_support_bil",
    "bank_treasury_current_demand_support_bil",
    "legacy_interest_support_bil",
    "projected_iorb_interest_basis_bil",
    "projected_iorb_current_demand_support_bil",
    "projected_on_rrp_interest_basis_bil",
    "projected_on_rrp_current_demand_support_bil",
    "projected_current_remittance_state_bil",
    "projected_current_remittance_demand_offset_bil",
    "projected_future_remittance_drag_bil",
    "projected_future_remittance_drag_demand_offset_bil",
    "gross_public_interest_current_demand_support_bil",
    "interest_income_tax_timing_drag_bil",
    "net_interest_before_fiscal_tga_offsets_bil",
    "fiscal_offset_bil",
    "tga_liquidity_offset_bil",
    "net_interest_after_fiscal_tga_offsets_bil",
    "replacement_delta_vs_legacy_interest_support_bil",
    "cbo_nominal_gdp_bil",
    "cbo_short_rate_pct",
    "reserve_balance_stock_gdp_share",
    "on_rrp_stock_gdp_share",
    "iorb_rate_spread_vs_cbo_short_rate_pct",
    "on_rrp_rate_spread_vs_cbo_short_rate_pct",
    "iorb_projection_status",
    "on_rrp_projection_status",
    "remittance_projection_status",
    "remittance_timing_treatment",
    "foreign_leakage_treatment",
    "tax_timing_treatment",
    "fiscal_tga_treatment",
    "composition_rule",
    "allowed_use",
    "blocked_use",
    "canonical_ratio_entry",
    "enters_main_ratio",
    "evidence_mode_enabled",
]

FED_FORECAST_SOURCE_FIELDS = [
    "forecast_fed_liability_source_row_id",
    "series_id",
    "series_label",
    "source_url",
    "source_cache_path",
    "observation_count",
    "latest_observation_date",
    "latest_observation_value",
    "latest_average_window_observation_count",
    "latest_average_value",
    "unit",
    "projection_use",
    "source_status",
]

RESIDUAL_NUMERATOR_SENSITIVITY_FIELDS = [
    "forecast_residual_numerator_sensitivity_row_id",
    "fiscal_year",
    "scenario_id",
    "baseline_scenario_id",
    "assumption_set",
    "source_vintage",
    "cbo_nominal_gdp_bil",
    "curve_effective_overlay_bp",
    "selected_moving_denominator_bil",
    "selected_delta_denominator_bil",
    "firm_liquid_asset_stock_gdp_share",
    "firm_liquid_asset_stock_source_status",
    "firm_cash_rate_path_yield_basis_bil",
    "firm_cash_attenuation_share",
    "firm_cash_attenuation_bil",
    "public_interest_private_recipient_cashflow_basis_bil",
    "public_interest_already_demand_converted_bil",
    "public_interest_residual_cashflow_basis_bil",
    "household_safe_asset_stock_share",
    "household_safe_asset_access_conditioner",
    "retail_safe_yield_pass_through_beta",
    "household_safe_yield_current_spend_share",
    "household_safe_yield_capture_bil",
    "deposit_mmf_incremental_access_share",
    "deposit_mmf_substitution_conditioner",
    "deposit_mmf_substitution_offset_bil",
    "credit_supply_drag_basis_bil",
    "deposit_mmf_substitution_drag_share",
    "deposit_mmf_substitution_drag_bil",
    "paired_deposit_mmf_net_sensitivity_bil",
    "total_residual_sensitivity_bil",
    "baseline_firm_cash_attenuation_bil",
    "delta_firm_cash_attenuation_vs_baseline_bil",
    "baseline_household_safe_yield_capture_bil",
    "delta_household_safe_yield_capture_vs_baseline_bil",
    "baseline_paired_deposit_mmf_net_sensitivity_bil",
    "delta_paired_deposit_mmf_net_sensitivity_vs_baseline_bil",
    "baseline_total_residual_sensitivity_bil",
    "delta_total_residual_sensitivity_vs_baseline_bil",
    "firm_cash_status",
    "safe_yield_status",
    "deposit_mmf_pairing_status",
    "denominator_overlap_status",
    "composition_rule",
    "allowed_use",
    "blocked_use",
    "canonical_ratio_entry",
    "enters_main_ratio",
    "evidence_mode_enabled",
]

RESIDUAL_SENSITIVITY_SOURCE_FIELDS = [
    "forecast_residual_sensitivity_source_row_id",
    "series_id",
    "series_label",
    "source_url",
    "source_cache_path",
    "observation_count",
    "latest_observation_date",
    "latest_observation_value",
    "latest_average_window_observation_count",
    "latest_average_value",
    "unit",
    "projection_use",
    "source_status",
]

FORECAST_COMPOSITION_SURFACE_FIELDS = [
    "forecast_composition_surface_row_id",
    "fiscal_year",
    "scenario_id",
    "baseline_scenario_id",
    "beta_path_id",
    "composition_case_id",
    "composition_case_label",
    "residual_assumption_set",
    "tdc_current_demand_support_bil",
    "direct_treasury_current_demand_support_bil",
    "bank_treasury_current_demand_support_bil",
    "legacy_interest_support_bil",
    "public_interest_net_support_bil",
    "interest_replacement_delta_bil",
    "residual_sensitivity_delta_bil",
    "first_forecast_n_bil",
    "composition_n_bil",
    "baseline_composition_n_bil",
    "delta_composition_n_vs_baseline_bil",
    "selected_moving_denominator_bil",
    "baseline_selected_moving_denominator_bil",
    "selected_delta_denominator_bil",
    "composition_ratewall_ratio",
    "baseline_composition_ratewall_ratio",
    "delta_composition_ratewall_ratio_vs_baseline",
    "wall_hit_under_composition",
    "composition_rule",
    "allowed_use",
    "blocked_use",
    "canonical_ratio_entry",
    "enters_main_ratio",
    "evidence_mode_enabled",
]

CENTRAL_FORECAST_SURFACE_FIELDS = [
    "central_forecast_surface_row_id",
    "fiscal_year",
    "scenario_id",
    "baseline_scenario_id",
    "central_beta_path_id",
    "central_composition_case_id",
    "central_n_bil",
    "central_moving_denominator_bil",
    "central_ratewall_ratio",
    "baseline_central_n_bil",
    "baseline_central_moving_denominator_bil",
    "baseline_central_ratewall_ratio",
    "delta_central_n_vs_baseline_bil",
    "delta_central_moving_denominator_vs_baseline_bil",
    "delta_central_ratewall_ratio_vs_baseline",
    "wall_hit_under_central_forecast",
    "first_forecast_ratewall_ratio",
    "delta_first_forecast_ratewall_ratio_vs_central",
    "residual_base_ratewall_ratio",
    "delta_residual_base_ratewall_ratio_vs_central",
    "residual_paired_ratewall_ratio",
    "delta_residual_paired_ratewall_ratio_vs_central",
    "latest_rolling_beta_ratewall_ratio",
    "delta_latest_rolling_beta_ratewall_ratio_vs_central",
    "pooled_full_beta_ratewall_ratio",
    "delta_pooled_full_beta_ratewall_ratio_vs_central",
    "central_choice_status",
    "sensitivity_rule",
    "allowed_use",
    "blocked_use",
    "canonical_ratio_entry",
    "enters_main_ratio",
    "evidence_mode_enabled",
]

CENTRAL_SCENARIO_INTERPRETATION_FIELDS = [
    "central_scenario_interpretation_row_id",
    "fiscal_year",
    "scenario_id",
    "baseline_scenario_id",
    "central_ratewall_ratio",
    "delta_central_ratewall_ratio_vs_baseline",
    "delta_central_n_vs_baseline_bil",
    "delta_central_moving_denominator_vs_baseline_bil",
    "numerator_only_delta_ratewall_ratio",
    "denominator_only_delta_ratewall_ratio",
    "primary_driver",
    "scenario_direction",
    "mechanism_summary",
    "sensitivity_min_ratewall_ratio",
    "sensitivity_max_ratewall_ratio",
    "sensitivity_width_ratewall_ratio",
    "largest_sensitivity_case",
    "largest_sensitivity_delta_vs_central",
    "allowed_use",
    "blocked_use",
    "canonical_ratio_entry",
    "enters_main_ratio",
    "evidence_mode_enabled",
]

FORECAST_SCENARIO_SUFFICIENCY_FIELDS = [
    "forecast_scenario_sufficiency_row_id",
    "scenario_id",
    "scenario_title",
    "scenario_axis",
    "configured_in_manifest",
    "run_in_suite",
    "in_central_surface",
    "provenance_kind",
    "tdc_route_taxonomy",
    "coverage_status",
    "sufficiency_decision",
    "next_model_action",
    "fy2036_delta_central_ratewall_ratio",
    "fy2036_primary_driver",
    "allowed_use",
    "blocked_use",
    "canonical_ratio_entry",
    "enters_main_ratio",
    "evidence_mode_enabled",
]

RESERVE_USER_TDC_ROUTE_HOLDER_TYPES = ("Banks", "Foreign", "CB", "FedInternal")
PRIVATE_DEPOSIT_USER_TDC_ROUTE_HOLDER_TYPES = ("Private",)

FRED_SERIES: Mapping[str, Mapping[str, str]] = {
    "WRBWFRBL": {
        "label": "Reserve balances with Federal Reserve Banks",
        "unit": "millions_of_dollars",
        "projection_use": "latest_13_observation_average_stock_to_gdp_share",
    },
    "IORB": {
        "label": "Interest rate on reserve balances",
        "unit": "percent",
        "projection_use": "latest_13_observation_average_rate_spread_vs_cbo_short_rate",
    },
    "RRPONTSYD": {
        "label": "ON RRP Treasury securities sold",
        "unit": "millions_of_dollars",
        "projection_use": "latest_13_observation_average_stock_to_gdp_share",
    },
    "RRPONTSYAWARD": {
        "label": "ON RRP award rate",
        "unit": "percent",
        "projection_use": "latest_13_observation_average_rate_spread_vs_cbo_short_rate",
    },
    "RESPPLLOPNWW": {
        "label": "Earnings remittances due to U.S. Treasury",
        "unit": "millions_of_dollars",
        "projection_use": "current_state_guard_only_not_annual_flow",
    },
    "WDTGAL": {
        "label": "U.S. Treasury General Account",
        "unit": "millions_of_dollars",
        "projection_use": "state_context_only_zero_baseline_without_tga_path",
    },
}

RESIDUAL_SENSITIVITY_FRED_SERIES: Mapping[str, Mapping[str, str]] = {
    "NCBCDCA": {
        "label": (
            "Nonfinancial corporate business checkable deposits and currency"
        ),
        "unit": "millions_of_dollars",
        "projection_use": "firm_liquid_asset_stock_component_latest_4q_average",
    },
    "TSDABSNNCB": {
        "label": "Nonfinancial corporate business time and savings deposits",
        "unit": "millions_of_dollars",
        "projection_use": "firm_liquid_asset_stock_component_latest_4q_average",
    },
    "TSABSNNCB": {
        "label": "Nonfinancial corporate business Treasury securities",
        "unit": "millions_of_dollars",
        "projection_use": "firm_liquid_asset_stock_component_latest_4q_average",
    },
    "BOGZ1FL103034000Q": {
        "label": "Nonfinancial corporate business money market fund shares",
        "unit": "millions_of_dollars",
        "projection_use": "firm_liquid_asset_stock_component_latest_4q_average",
    },
    "SRPSABSNNCB": {
        "label": (
            "Nonfinancial corporate business security repurchase agreements"
        ),
        "unit": "millions_of_dollars",
        "projection_use": "firm_liquid_asset_stock_component_latest_4q_average",
    },
}

FRED_GRAPH_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
DEFAULT_FED_SOURCE_CACHE_DIR = Path(
    "var/preliminary_scenario_results/forecast_10y/source_cache/fred"
)
CBO_ECONOMIC_PROJECTIONS_XLSX = Path(
    "data/raw/cbo/51135-2026-02-Economic-Projections.xlsx"
)
CBO_BUDGET_PROJECTIONS_XLSX = Path(
    "data/raw/cbo/51118-2026-02-Budget-Projections.xlsx"
)


class ForecastModelReadoutError(ValueError):
    """Raised when the ten-year forecast readout cannot be assembled."""


@dataclass(frozen=True)
class _SuiteFiles:
    root: Path
    artifact: ArtifactManifestView | None


_BETA_PROFILE_BY_ID = {
    label: (beta, source_status)
    for label, beta, source_status in BETA_CHI_ROBUSTNESS_BETA_PROFILES
}

TIMED_BETA_PATHS = (
    {
        "beta_path_id": "normal_forward_constant",
        "beta_path_label": "Normal-forward beta for every forecast year",
        "transition_year": "",
        "early_beta": "normal_forward",
        "late_beta": "normal_forward",
    },
    {
        "beta_path_id": "latest_rolling_from_fy2032",
        "beta_path_label": "Normal-forward beta through FY2031, latest rolling from FY2032",
        "transition_year": "2032",
        "early_beta": "normal_forward",
        "late_beta": "latest_rolling_persistence",
    },
    {
        "beta_path_id": "pooled_full_from_fy2032_review",
        "beta_path_label": "Normal-forward beta through FY2031, pooled full-sample from FY2032",
        "transition_year": "2032",
        "early_beta": "normal_forward",
        "late_beta": "pooled_full_sample",
    },
)

ACTIVE_FORECAST_CHANNELS = (
    (
        "tdc_ex_overlap_current_demand_support",
        "TDC ex-overlap support",
        "included_in_first_10y_forecast",
        "TDCSim exports ex-overlap TDC flows and RateWall applies beta times chi once.",
    ),
    (
        "direct_treasury_interest_support",
        "Direct Treasury interest support",
        "included_in_first_10y_forecast",
        "TDCSim exports the direct Treasury interest support used in the first numerator.",
    ),
    (
        "bank_treasury_interest_support",
        "Bank Treasury interest support",
        "included_in_first_10y_forecast",
        "TDCSim exports the bank Treasury interest support used in the first numerator.",
    ),
)

DEFERRED_FORECAST_CHANNELS = (
    "net_interest_after_fiscal_tga_offsets",
    "current_remittance_demand_offset",
    "future_remittance_drag_demand_offset",
    "iorb_recipient_demand_channel",
    "on_rrp_recipient_demand_channel",
    "fiscal_offset",
    "tga_liquidity_offset",
    "foreign_treasury_holder_leakage_drag",
    "interest_income_tax_timing_drag",
    "firm_cash_attenuation",
    "safe_asset_allocation_offset",
    "safe_asset_allocation_drag",
    "zero_interest_credit_attenuation",
    "household_safe_yield_capture",
    "deposit_mmf_substitution_offset",
    "deposit_mmf_substitution_drag",
    "firm_liquid_asset_cushion",
    "firm_rollover_pressure_drag",
)

DEFERRED_CHANNEL_FINAL_TREATMENT: Mapping[str, Mapping[str, str]] = {
    "net_interest_after_fiscal_tga_offsets": {
        "classification": "included_as_public_interest_replacement_block",
        "reason": (
            "New forecast public-interest net block replaces direct plus bank "
            "interest rows for final accounting."
        ),
        "next_model_requirement": "use_net_block_in_final_n_composition_never_add_both",
        "blocked_use": "add_on_top_of_direct_and_bank_interest_rows",
        "public_interest_block_role": "block_output_standalone_final_n_term",
    },
    "current_remittance_demand_offset": {
        "classification": "integrated_in_public_interest_net_block_with_state_guard",
        "reason": (
            "Included inside the net block only when a clean annual positive "
            "remittance projection exists; H.4.1 state is memo-only otherwise."
        ),
        "next_model_requirement": "extract_clean_cbo_annual_remittance_level_if_available",
        "blocked_use": "convert_negative_deferred_asset_or_weekly_state_to_annual_support",
        "public_interest_block_role": "signed_remittance_timing_subchannel",
    },
    "future_remittance_drag_demand_offset": {
        "classification": "integrated_in_public_interest_net_block_with_state_guard",
        "reason": (
            "Carried inside the net block as a signed future-remittance timing "
            "subchannel; forecast remains zero without a clean annual drag path."
        ),
        "next_model_requirement": (
            "carry_signed_future_remittance_drag_inside_public_interest_block"
        ),
        "blocked_use": "drop_future_remittance_drag_or_treat_as_standalone_support",
        "public_interest_block_role": "signed_future_remittance_timing_subchannel",
    },
    "iorb_recipient_demand_channel": {
        "classification": "integrated_in_public_interest_net_block",
        "reason": "Projected from reserve balances, CBO short rates, and the live IORB demand-share prior.",
        "next_model_requirement": "keep_fed_liability_source_cache_current",
        "blocked_use": "standalone_additive_channel_or_bank_incidence_claim",
        "public_interest_block_role": "fed_liability_interest_input",
    },
    "on_rrp_recipient_demand_channel": {
        "classification": "integrated_in_public_interest_net_block",
        "reason": "Projected from ON RRP stock, CBO short rates, and the live ON RRP demand-share prior.",
        "next_model_requirement": "zero_mechanically_when_on_rrp_stock_is_zero",
        "blocked_use": "reuse_tdcsim_mmf_route_coefficient_as_ratewall_beta",
        "public_interest_block_role": "fed_liability_interest_input",
    },
    "fiscal_offset": {
        "classification": "integrated_in_public_interest_net_block_absorber",
        "reason": "Applied after tax timing inside the public-interest net block.",
        "next_model_requirement": "keep_assumption_mode_absorber_label",
        "blocked_use": "fiscal_reaction_estimate_or_tax_output",
        "public_interest_block_role": "absorber",
    },
    "tga_liquidity_offset": {
        "classification": "integrated_in_public_interest_net_block_absorber",
        "reason": "Applied after tax timing inside the public-interest net block.",
        "next_model_requirement": "replace_scalar_with_explicit_tga_path_only_if_available",
        "blocked_use": "permanent_liquidity_claim_without_tga_path",
        "public_interest_block_role": "absorber",
    },
    "foreign_treasury_holder_leakage_drag": {
        "classification": "replaced_by_current_tdcsim_holder_filter",
        "reason": (
            "The current direct and bank Treasury interest bases are already "
            "holder-filtered; a second foreign haircut would double subtract."
        ),
        "next_model_requirement": "apply_once_only_if_all_holder_treasury_basis_returns",
        "blocked_use": "second_foreign_haircut_on_tdcsim_filtered_basis",
    },
    "interest_income_tax_timing_drag": {
        "classification": "integrated_in_public_interest_net_block_absorber",
        "reason": "Applied to positive gross public-interest support before fiscal and TGA absorbers.",
        "next_model_requirement": "keep_timing_haircut_not_tax_output_label",
        "blocked_use": "tax_incidence_or_welfare_claim",
        "public_interest_block_role": "absorber",
    },
    "firm_cash_attenuation": {
        "classification": "projection_required_as_bounded_sensitivity",
        "reason": "Material enough to keep, but still weak-link Assumption Mode rather than central evidence.",
        "next_model_requirement": "project_from_cbo_gdp_rate_path_and_z1_firm_liquid_asset_context",
        "blocked_use": "unqualified_central_claim_or_add_with_firm_cushion",
    },
    "safe_asset_allocation_offset": {
        "classification": "obsolete_due_to_overlap_guard",
        "reason": "Superseded by explicit recipient and residual safe-yield channels unless rebuilt on a disjoint basis.",
        "next_model_requirement": "keep_zero_when_same_basis_channels_active",
        "blocked_use": "same_basis_additive_safe_asset_offset",
    },
    "safe_asset_allocation_drag": {
        "classification": "sidecar_until_disjoint_basis_exists",
        "reason": "Do not subtract from final N without a disjoint safe-asset drag basis.",
        "next_model_requirement": "build_disjoint_basis_or_leave_sidecar",
        "blocked_use": "conservative_double_count_subtraction",
    },
    "zero_interest_credit_attenuation": {
        "classification": "explicit_limitation_or_minor_sensitivity",
        "reason": "Low-materiality credit-side adjacent channel; not central without stock and duration data.",
        "next_model_requirement": "keep_out_of_central_forecast_until_material_stock_path_exists",
        "blocked_use": "measured_credit_relief_claim_without_source_stock",
    },
    "household_safe_yield_capture": {
        "classification": "projection_required_as_bounded_residual_sensitivity",
        "reason": "May enter only on a residual domestic safe-yield basis after direct recipient conversion.",
        "next_model_requirement": "build_residual_basis_and_no_double_count_test",
        "blocked_use": "same_cashflow_dollar_converted_twice",
    },
    "deposit_mmf_substitution_offset": {
        "classification": "projection_required_as_paired_bounded_sensitivity",
        "reason": "Only admissible with the matching drag row and residual basis.",
        "next_model_requirement": "build_paired_offset_drag_sensitivity",
        "blocked_use": "unpaired_mmf_offset_or_mmf_route_beta_confusion",
    },
    "deposit_mmf_substitution_drag": {
        "classification": "projection_required_as_paired_bounded_sensitivity",
        "reason": "Only admissible with the matching offset row and denominator-overlap guard.",
        "next_model_requirement": "build_paired_offset_drag_sensitivity",
        "blocked_use": "credit_drag_double_count_against_moving_D",
    },
    "firm_liquid_asset_cushion": {
        "classification": "diagnostic_or_replacement_only",
        "reason": "Must not enter together with firm cash attenuation on the same firm-liquid-asset basis.",
        "next_model_requirement": "use_only_if_firm_cash_is_demoted",
        "blocked_use": "additive_firm_cash_plus_firm_cushion",
    },
    "firm_rollover_pressure_drag": {
        "classification": "denominator_sidecar_not_numerator",
        "reason": "Belongs with borrowing/credit drag interpretation, not current-demand support.",
        "next_model_requirement": "keep_as_moving_D_or_credit_sidecar",
        "blocked_use": "negative_numerator_addition_from_denominator_drag_basis",
    },
}

REMAINING_NUMERATOR_CHANNEL_PLAN: Mapping[str, Mapping[str, str]] = {
    "safe_asset_allocation_drag": {
        "final_central_status": "not_admitted_pending_disjoint_basis",
        "materiality_tier": "possibly_material_but_currently_double_count_prone",
        "forecast_route": (
            "assumption_bounds_only_after_building_a_residual_safe_asset_drag_basis"
        ),
        "assumption_basis": (
            "literature_calibrated_wealth_or_liquidity_drag_on_residual_domestic_safe_asset_holdings"
        ),
        "calibration_need": (
            "source_backed_wealth_effect_or_liquidity_discount_by_duration_plus_residual_holder_stock_path"
        ),
        "double_count_guard": (
            "must_not_overlap_household_safe_yield_capture_public_interest_block_or_moving_D"
        ),
        "admission_test": (
            "nonzero_only_if_the_cashflow_or_stock_basis_is_disjoint_and_the_bound_is_calibrated"
        ),
        "next_model_action": (
            "seek_literature_calibrated_bound_or_keep_as_sidecar_limitation"
        ),
        "blocked_use": "same_basis_safe_asset_drag_subtracted_from_central_N",
    },
    "zero_interest_credit_attenuation": {
        "final_central_status": "minor_sensitivity_only_unless_stock_path_is_material",
        "materiality_tier": (
            "bnpl_likely_low_total_zero_low_apr_promotional_credit_unproven"
        ),
        "forecast_route": (
            "product_specific_materiality_screen_then_minor_sensitivity_if_source_backed"
        ),
        "assumption_basis": (
            "bnpl_pay_in_4_longer_bnpl_credit_card_promo_apr_and_deferred_interest_split"
        ),
        "calibration_need": (
            "average_outstanding_zero_low_apr_stock_duration_rate_wedge_and_cashflow_response"
        ),
        "double_count_guard": (
            "must_not_relabel_moving_D_credit_supply_effect_as_current_demand_N"
        ),
        "admission_test": (
            "enter_only_as_minor_sensitivity_when_product_stock_path_and_bound_are_present"
        ),
        "next_model_action": (
            "run_product_specific_zero_low_apr_credit_materiality_screen"
        ),
        "blocked_use": "measured_credit_relief_claim_without_source_stock",
    },
    "firm_liquid_asset_cushion": {
        "final_central_status": "replacement_candidate_not_additive",
        "materiality_tier": "material_only_as_alternative_to_firm_cash_attenuation",
        "forecast_route": (
            "replacement_scenario_for_firm_cash_attenuation_using_z1_firm_liquid_asset_context"
        ),
        "assumption_basis": (
            "cash_buffer_or_liquid_asset_cushion_share_applied_to_the_same_firm_liquid_asset_basis"
        ),
        "calibration_need": (
            "evidence_for_cushion_share_or_a_rule_that_demotes_firm_cash_attenuation"
        ),
        "double_count_guard": (
            "cannot_enter_together_with_firm_cash_attenuation_on_same_asset_basis"
        ),
        "admission_test": (
            "central_N_can_use_either_firm_cash_attenuation_or_cushion_replacement_not_both"
        ),
        "next_model_action": (
            "keep_as_replacement_case_unless_firm_cash_attenuation_is_demoted"
        ),
        "blocked_use": "additive_firm_cash_plus_firm_cushion",
    },
    "firm_rollover_pressure_drag": {
        "final_central_status": "not_a_current_numerator_channel_without_new_credit_model",
        "materiality_tier": "possibly_material_for_credit_conditions_not_current_N",
        "forecast_route": (
            "denominator_or_credit_sidecar_until_a_disjoint_current_spending_drag_is_defined"
        ),
        "assumption_basis": (
            "firm_debt_maturity_wall_and_refinancing_cost_channel_if_modeled_separately"
        ),
        "calibration_need": (
            "firm_debt_maturity_stock_path_refinancing_spread_and_spending_response"
        ),
        "double_count_guard": (
            "must_not_double_count_rate_scenarios_already_moving_D"
        ),
        "admission_test": (
            "do_not_enter_central_N_without_a_current_spending_basis_not_already_in_D"
        ),
        "next_model_action": (
            "park_for_later_denominator_or_credit_model_not_current_forecast_N"
        ),
        "blocked_use": "negative_numerator_addition_from_denominator_drag_basis",
    },
}

ZERO_LOW_APR_CREDIT_SCREEN_ROWS: tuple[Mapping[str, str], ...] = (
    {
        "product_segment": "bnpl_pay_in_4_average_outstanding",
        "source_vintage": "richmond_fed_2026",
        "source_url": "https://www.richmondfed.org/publications/research/economic_brief/2026/eb_26-05",
        "source_metric": "2025_bnpl_transaction_value_70bn_implied_average_outstanding_3_02bn",
        "annual_originations_bil": "70",
        "outstanding_stock_bil": "3.02",
        "zero_low_apr_share": "1",
        "duration_months": "short_pay_in_4_outstanding_stock_already_duration_adjusted",
        "candidate_rate_wedge_pct": "21.52",
        "screen_status": "source_backed_small_outstanding_upper_bound",
        "central_n_treatment": "not_in_central_n",
        "materiality_decision": "low_materiality_for_pay_in_4_bnpl_only",
        "next_model_action": "no_central_entry_keep_as_context",
        "blocked_use": "scale_from_originations_or_generalize_to_all_zero_low_apr_credit",
    },
    {
        "product_segment": "broader_bnpl_zero_apr_originations",
        "source_vintage": "federal_reserve_feds_notes_2026",
        "source_url": "https://www.federalreserve.gov/econres/notes/feds-notes/buy-now-pay-later-beyond-pay-in-4-a-comprehensive-product-overview-20260605.html",
        "source_metric": "2025_bnpl_originations_near_160bn_more_than_60pct_zero_apr",
        "annual_originations_bil": "160",
        "outstanding_stock_bil": "",
        "zero_low_apr_share": "0.60",
        "duration_months": "missing_product_outstanding_duration",
        "candidate_rate_wedge_pct": "21.52",
        "screen_status": "originations_only_not_stock_path",
        "central_n_treatment": "not_in_central_n",
        "materiality_decision": "cannot_size_current_support_from_originations",
        "next_model_action": "requires_product_outstanding_stock_before_any_sensitivity",
        "blocked_use": "originations_as_current_outstanding_stock",
    },
    {
        "product_segment": "credit_card_introductory_promo_apr_balances",
        "source_vintage": "boston_philadelphia_fed_2018_2019_and_richmond_fed_2025",
        "source_url": "https://www.bostonfed.org/news-and-events/news/2024/03/zero-percent-apr-credit-cards-are-everywhere-but-how-does-consumer-behavior-impact-their-popularity.aspx",
        "source_metric": "historical_25pct_card_debt_intro_offer_mostly_zero_apr_avg_9mo_reset_16pp_contextualized_against_1_23tn_card_debt",
        "annual_originations_bil": "",
        "outstanding_stock_bil": "307.5",
        "zero_low_apr_share": "0.25",
        "duration_months": "9",
        "candidate_rate_wedge_pct": "16",
        "screen_status": "potentially_material_but_historical_share_not_current_path",
        "central_n_treatment": "not_in_central_n",
        "materiality_decision": "materiality_risk_requires_current_product_stock_path",
        "next_model_action": "seek_current_promo_apr_outstanding_or_keep_as_limitation",
        "blocked_use": "apply_2018_2019_promo_share_as_current_forecast_without_source",
    },
    {
        "product_segment": "deferred_interest_retail_credit",
        "source_vintage": "no_current_source_in_repo",
        "source_url": "",
        "source_metric": "product_exists_but_current_outstanding_stock_missing",
        "annual_originations_bil": "",
        "outstanding_stock_bil": "",
        "zero_low_apr_share": "",
        "duration_months": "missing",
        "candidate_rate_wedge_pct": "",
        "screen_status": "missing_required_stock_and_duration",
        "central_n_treatment": "not_in_central_n",
        "materiality_decision": "park_until_source_stock_exists",
        "next_model_action": "do_not_model_without_source_outstanding_stock",
        "blocked_use": "assume_materiality_from_product_existence",
    },
)


def timed_beta_path_rows_from_directory(
    suite_dir: str | Path = DEFAULT_FORECAST_READOUT_SUITE_DIR,
) -> list[dict[str, str]]:
    """Build timed-beta forecast rows from the active model-summary surface."""

    files = _suite_files(suite_dir)
    effect_rows = _read_csv(files, "ratewall_tdcsim_cbo_scenario_effect.csv")
    scenario_config_dir = Path(suite_dir) / "scenarios"
    return timed_beta_path_rows(
        effect_rows=effect_rows,
        summary_rows=tdcsim_cbo_model_scenario_summary_rows(effect_rows),
        synthesis_rows=tdcsim_cbo_model_scenario_interpretation_synthesis_rows(
            effect_rows,
            scenario_config_dir=scenario_config_dir,
        ),
        materiality_rows=tdcsim_cbo_model_scenario_materiality_classification_rows(
            effect_rows,
            scenario_config_dir=scenario_config_dir,
        ),
    )


def forecast_scenario_sufficiency_rows_from_directory(
    suite_dir: str | Path = DEFAULT_FORECAST_READOUT_SUITE_DIR,
    *,
    central_interpretation_rows: Iterable[Mapping[str, str]],
) -> list[dict[str, str]]:
    """Review which scenario axes are covered by the current forecast surface."""

    files = _suite_files(suite_dir)
    return forecast_scenario_sufficiency_rows(
        effect_rows=_read_csv(files, "ratewall_tdcsim_cbo_scenario_effect.csv"),
        scenario_config_rows=_scenario_config_review_rows(files),
        central_interpretation_rows=central_interpretation_rows,
    )


def refresh_fed_forecast_source_cache(
    source_cache_dir: str | Path = DEFAULT_FED_SOURCE_CACHE_DIR,
) -> list[Path]:
    """Download the small official FRED CSVs used by the forecast source rows."""

    return _refresh_fred_source_cache(FRED_SERIES, source_cache_dir)


def refresh_residual_sensitivity_source_cache(
    source_cache_dir: str | Path = DEFAULT_FED_SOURCE_CACHE_DIR,
) -> list[Path]:
    """Download the small official FRED CSVs used by residual sensitivities."""

    return _refresh_fred_source_cache(RESIDUAL_SENSITIVITY_FRED_SERIES, source_cache_dir)


def _refresh_fred_source_cache(
    series: Mapping[str, Mapping[str, str]],
    source_cache_dir: str | Path,
) -> list[Path]:
    """Download FRED CSVs for one compact source set."""

    out = Path(source_cache_dir)
    out.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for series_id in series:
        path = out / f"{series_id}.csv"
        urlretrieve(FRED_GRAPH_URL.format(series_id=series_id), path)
        paths.append(path)
    return paths


def forecast_public_interest_net_block_rows_from_directory(
    suite_dir: str | Path = DEFAULT_FORECAST_READOUT_SUITE_DIR,
    *,
    source_cache_dir: str | Path | None = DEFAULT_FED_SOURCE_CACHE_DIR,
    assumption_set_name: str = "literature_calibrated_base",
    cbo_economic_workbook: str | Path = CBO_ECONOMIC_PROJECTIONS_XLSX,
    cbo_budget_workbook: str | Path = CBO_BUDGET_PROJECTIONS_XLSX,
    source_as_of_date: date | None = None,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Build the projected public-interest net block for the 10-year readout."""

    files = _suite_files(suite_dir)
    effect_rows = _read_csv(files, "ratewall_tdcsim_cbo_scenario_effect.csv")
    return forecast_public_interest_net_block_rows(
        effect_rows=effect_rows,
        cbo_macro_rows=cbo_fiscal_macro_rows(cbo_economic_workbook),
        fed_source_rows=fed_forecast_source_rows(
            source_cache_dir,
            as_of_date=source_as_of_date,
        )
        if source_cache_dir is not None
        else [],
        assumption_set_name=assumption_set_name,
        remittance_projection_rows=cbo_remittance_projection_rows(cbo_budget_workbook),
    )


def forecast_residual_numerator_sensitivity_rows_from_directory(
    suite_dir: str | Path = DEFAULT_FORECAST_READOUT_SUITE_DIR,
    *,
    source_cache_dir: str | Path | None = DEFAULT_FED_SOURCE_CACHE_DIR,
    assumption_set_names: Sequence[str] = (
        "literature_calibrated_base",
        "assumption_mode_deposit_mmf_paired_entry",
    ),
    cbo_economic_workbook: str | Path = CBO_ECONOMIC_PROJECTIONS_XLSX,
    source_as_of_date: date | None = None,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Build forecast rows for residual numerator sensitivities."""

    files = _suite_files(suite_dir)
    residual_source_rows = (
        residual_sensitivity_source_rows(source_cache_dir, as_of_date=source_as_of_date)
        if source_cache_dir is not None
        else []
    )
    effect_rows = _read_csv(files, "ratewall_tdcsim_cbo_scenario_effect.csv")
    return forecast_residual_numerator_sensitivity_rows(
        effect_rows=effect_rows,
        synthesis_rows=tdcsim_cbo_model_scenario_interpretation_synthesis_rows(
            effect_rows,
            scenario_config_dir=Path(suite_dir) / "scenarios",
        ),
        public_interest_rows=forecast_public_interest_net_block_rows_from_directory(
            suite_dir,
            source_cache_dir=source_cache_dir,
            cbo_economic_workbook=cbo_economic_workbook,
            source_as_of_date=source_as_of_date,
        )[0],
        cbo_macro_rows=cbo_fiscal_macro_rows(cbo_economic_workbook),
        residual_source_rows=residual_source_rows,
        assumption_set_names=assumption_set_names,
    )


def forecast_composition_surface_rows_from_directory(
    suite_dir: str | Path = DEFAULT_FORECAST_READOUT_SUITE_DIR,
    *,
    source_cache_dir: str | Path | None = DEFAULT_FED_SOURCE_CACHE_DIR,
    cbo_economic_workbook: str | Path = CBO_ECONOMIC_PROJECTIONS_XLSX,
    source_as_of_date: date | None = None,
) -> list[dict[str, str]]:
    """Build the final forecast-composition comparison surface."""

    timed_rows = timed_beta_path_rows_from_directory(suite_dir)
    public_interest_rows, _fed_source_rows = (
        forecast_public_interest_net_block_rows_from_directory(
            suite_dir,
            source_cache_dir=source_cache_dir,
            cbo_economic_workbook=cbo_economic_workbook,
            source_as_of_date=source_as_of_date,
        )
    )
    residual_rows, _residual_source_rows = (
        forecast_residual_numerator_sensitivity_rows_from_directory(
            suite_dir,
            source_cache_dir=source_cache_dir,
            cbo_economic_workbook=cbo_economic_workbook,
            source_as_of_date=source_as_of_date,
        )
    )
    return forecast_composition_surface_rows(
        timed_beta_rows=timed_rows,
        public_interest_rows=public_interest_rows,
        residual_sensitivity_rows=residual_rows,
    )


def central_forecast_surface_rows_from_directory(
    suite_dir: str | Path = DEFAULT_FORECAST_READOUT_SUITE_DIR,
    *,
    source_cache_dir: str | Path | None = DEFAULT_FED_SOURCE_CACHE_DIR,
    cbo_economic_workbook: str | Path = CBO_ECONOMIC_PROJECTIONS_XLSX,
    source_as_of_date: date | None = None,
) -> list[dict[str, str]]:
    """Build the selected central forecast surface."""

    return central_forecast_surface_rows(
        forecast_composition_surface_rows_from_directory(
            suite_dir,
            source_cache_dir=source_cache_dir,
            cbo_economic_workbook=cbo_economic_workbook,
            source_as_of_date=source_as_of_date,
        )
    )


def forecast_public_interest_net_block_rows(
    *,
    effect_rows: Iterable[Mapping[str, str]],
    cbo_macro_rows: Iterable[Mapping[str, str]],
    fed_source_rows: Iterable[Mapping[str, str]],
    assumption_set_name: str = "literature_calibrated_base",
    remittance_projection_rows: Iterable[Mapping[str, str]] = (),
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Restage direct/bank interest into one net public-interest block."""

    assumption = _assumption_set(assumption_set_name)
    macro_by_year = {row["fiscal_year"]: row for row in cbo_macro_rows}
    source_list = list(fed_source_rows)
    source_by_series = {row["series_id"]: row for row in source_list}
    remittance_by_year = {
        row["fiscal_year"]: row for row in remittance_projection_rows
    }
    if "WRBWFRBL" in source_by_series:
        reserve_share = _stock_gdp_share(
            source_by_series["WRBWFRBL"],
            macro_by_year,
            "Reserve balances",
        )
        iorb_stock_status = "source_backed_latest_reserve_balance_gdp_share_held_constant"
    else:
        reserve_share = Decimal("0")
        iorb_stock_status = "explicit_zero_missing_reserve_balance_source"
    if "RRPONTSYD" in source_by_series:
        on_rrp_share = _stock_gdp_share(
            source_by_series["RRPONTSYD"],
            macro_by_year,
            "ON RRP",
        )
        on_rrp_stock_status = "source_backed_latest_on_rrp_gdp_share_held_constant"
    else:
        on_rrp_share = Decimal("0")
        on_rrp_stock_status = "explicit_zero_missing_on_rrp_source"
    iorb_spread, iorb_rate_status = _rate_spread(
        source_by_series.get("IORB"),
        macro_by_year,
        missing_status="missing_iorb_rate_source_uses_cbo_short_rate",
    )
    on_rrp_spread, on_rrp_rate_status = _rate_spread(
        source_by_series.get("RRPONTSYAWARD"),
        macro_by_year,
        missing_status="missing_on_rrp_award_rate_source_uses_cbo_short_rate",
    )
    remittance_source = source_by_series.get("RESPPLLOPNWW")
    remittance_status = (
        "cbo_remittance_projection_baseline_budget_context_not_numerator"
        if remittance_by_year
        else (
            "explicit_zero_baseline_no_cbo_annual_remittance_projection_"
            "h41_state_memo_only"
        )
    )
    if remittance_source is None and not remittance_by_year:
        remittance_status = "explicit_zero_missing_h41_remittance_state_source"

    rows: list[dict[str, str]] = []
    for effect in effect_rows:
        fiscal_year = effect["fiscal_year"]
        macro = macro_by_year.get(fiscal_year)
        if macro is None:
            raise ForecastModelReadoutError(
                f"missing CBO macro row for fiscal year {fiscal_year}"
            )
        gdp = _decimal(macro["cbo_nominal_gdp_bil"])
        short_rate = _decimal(macro["cbo_short_rate_pct"])
        direct = _decimal(effect["direct_treasury_current_demand_support_bil"])
        bank = _decimal(effect["bank_treasury_current_demand_support_bil"])
        legacy_interest_support = direct + bank
        iorb_basis = gdp * reserve_share * max(short_rate + iorb_spread, Decimal("0")) / Decimal("100")
        on_rrp_basis = gdp * on_rrp_share * max(short_rate + on_rrp_spread, Decimal("0")) / Decimal("100")
        iorb_support = (
            iorb_basis
            * _decimal(assumption.iorb_pass_through_scale)
            * _decimal(assumption.iorb_recipient_demand_share)
        )
        on_rrp_support = (
            on_rrp_basis
            * _decimal(assumption.on_rrp_pass_through_scale)
            * _decimal(assumption.on_rrp_recipient_demand_share)
        )
        remittance_projection = remittance_by_year.get(fiscal_year)
        annual_remittance_state = (
            max(
                _decimal(remittance_projection["cbo_federal_reserve_remittance_bil"]),
                Decimal("0"),
            )
            if remittance_projection is not None
            else Decimal("0")
        )
        remittance_support = Decimal("0")
        future_remittance_drag = Decimal("0")
        future_remittance_drag_offset = -(
            future_remittance_drag
            * _decimal(assumption.future_remittance_drag_timing_share)
            * _decimal(assumption.future_remittance_drag_demand_share)
        )
        gross = (
            legacy_interest_support
            + iorb_support
            + on_rrp_support
            + remittance_support
            + future_remittance_drag_offset
        )
        tax_drag = max(gross, Decimal("0")) * _decimal(
            assumption.interest_income_tax_timing_leakage_share
        )
        pre_fiscal = max(gross - tax_drag, Decimal("0"))
        fiscal_offset = pre_fiscal * _decimal(assumption.fiscal_offset_share)
        tga_offset = pre_fiscal * _decimal(assumption.tga_liquidity_offset_share)
        net = max(pre_fiscal - fiscal_offset - tga_offset, Decimal("0"))
        rows.append(
            {
                "forecast_public_interest_net_block_row_id": (
                    "forecast_public_interest_net_block::"
                    f"{fiscal_year}::{effect['scenario_id']}"
                ),
                "fiscal_year": fiscal_year,
                "scenario_id": effect["scenario_id"],
                "baseline_scenario_id": effect["baseline_scenario_id"],
                "assumption_set": assumption.name,
                "source_vintage": "cbo_2026_02_fred_latest_cache",
                "direct_treasury_current_demand_support_bil": _fmt(direct),
                "bank_treasury_current_demand_support_bil": _fmt(bank),
                "legacy_interest_support_bil": _fmt(legacy_interest_support),
                "projected_iorb_interest_basis_bil": _fmt(iorb_basis),
                "projected_iorb_current_demand_support_bil": _fmt(iorb_support),
                "projected_on_rrp_interest_basis_bil": _fmt(on_rrp_basis),
                "projected_on_rrp_current_demand_support_bil": _fmt(on_rrp_support),
                "projected_current_remittance_state_bil": _fmt(annual_remittance_state),
                "projected_current_remittance_demand_offset_bil": _fmt(
                    remittance_support
                ),
                "projected_future_remittance_drag_bil": _fmt(
                    future_remittance_drag
                ),
                "projected_future_remittance_drag_demand_offset_bil": _fmt(
                    future_remittance_drag_offset
                ),
                "gross_public_interest_current_demand_support_bil": _fmt(gross),
                "interest_income_tax_timing_drag_bil": _fmt(tax_drag),
                "net_interest_before_fiscal_tga_offsets_bil": _fmt(pre_fiscal),
                "fiscal_offset_bil": _fmt(fiscal_offset),
                "tga_liquidity_offset_bil": _fmt(tga_offset),
                "net_interest_after_fiscal_tga_offsets_bil": _fmt(net),
                "replacement_delta_vs_legacy_interest_support_bil": _fmt(
                    net - legacy_interest_support
                ),
                "cbo_nominal_gdp_bil": _fmt(gdp),
                "cbo_short_rate_pct": _fmt(short_rate),
                "reserve_balance_stock_gdp_share": _fmt(reserve_share),
                "on_rrp_stock_gdp_share": _fmt(on_rrp_share),
                "iorb_rate_spread_vs_cbo_short_rate_pct": _fmt(iorb_spread),
                "on_rrp_rate_spread_vs_cbo_short_rate_pct": _fmt(on_rrp_spread),
                "iorb_projection_status": f"{iorb_stock_status};{iorb_rate_status}",
                "on_rrp_projection_status": (
                    f"{on_rrp_stock_status};{on_rrp_rate_status}"
                ),
                "remittance_projection_status": remittance_status,
                "remittance_timing_treatment": (
                    "cbo_baseline_remittance_budget_context_only;"
                    "no_current_n_support_without_ex_overlap_delta_model"
                ),
                "foreign_leakage_treatment": (
                    "replaced_by_tdcsim_domestic_nonbank_and_bank_holder_basis_"
                    "no_second_foreign_haircut"
                ),
                "tax_timing_treatment": (
                    "assumption_mode_absorber_applied_after_gross_interest_support"
                ),
                "fiscal_tga_treatment": (
                    "assumption_mode_absorbers_applied_after_tax_timing"
                ),
                "composition_rule": (
                    "final_interest_block_replaces_legacy_direct_plus_bank_rows_"
                    "never_add_both"
                ),
                "allowed_use": "forecast_numerator_replacement_block_design",
                "blocked_use": (
                    "add_on_top_of_direct_and_bank_interest_rows;tax_output;"
                    "fiscal_reaction_estimate;incidence_or_welfare_claim;"
                    "canonical_headline_promotion"
                ),
                "canonical_ratio_entry": "false",
                "enters_main_ratio": "false",
                "evidence_mode_enabled": "false",
            }
        )
    return sorted(rows, key=lambda row: (int(row["fiscal_year"]), row["scenario_id"])), source_list


def forecast_residual_numerator_sensitivity_rows(
    *,
    effect_rows: Iterable[Mapping[str, str]],
    synthesis_rows: Iterable[Mapping[str, str]],
    public_interest_rows: Iterable[Mapping[str, str]],
    cbo_macro_rows: Iterable[Mapping[str, str]],
    residual_source_rows: Iterable[Mapping[str, str]],
    assumption_set_names: Sequence[str] = (
        "literature_calibrated_base",
        "assumption_mode_deposit_mmf_paired_entry",
    ),
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Measure residual channels without promoting overlap-prone dollars."""

    effects = _by_key(effect_rows, "scenario effect")
    synthesis = _by_key(synthesis_rows, "model synthesis")
    public_interest = _by_key(public_interest_rows, "public-interest net block")
    macro_by_year = {row["fiscal_year"]: row for row in cbo_macro_rows}
    source_list = list(residual_source_rows)
    firm_stock_share, firm_stock_status = _firm_liquid_asset_stock_share(
        source_list,
        macro_by_year,
        _assumption_set(assumption_set_names[0]),
    )
    rows: list[dict[str, str]] = []
    for assumption_name in assumption_set_names:
        assumption = _assumption_set(assumption_name)
        for key, syn in synthesis.items():
            effect = _required(effects, key, "scenario effect")
            year = effect["fiscal_year"]
            macro = macro_by_year.get(year)
            if macro is None:
                raise ForecastModelReadoutError(
                    f"missing CBO macro row for fiscal year {year}"
                )
            interest = _required(public_interest, key, "public-interest net block")
            gdp = _decimal(macro["cbo_nominal_gdp_bil"])
            curve_bp = _decimal(syn["curve_effective_overlay_bp"])
            firm_cash_basis = gdp * firm_stock_share * curve_bp / Decimal("10000")
            firm_cash = firm_cash_basis * _decimal(
                assumption.firm_cash_attenuation_share
            )
            private_basis = (
                _decimal(effect["direct_treasury_current_demand_support_bil"])
                + _decimal(effect["bank_treasury_current_demand_support_bil"])
                + _decimal(interest["projected_iorb_interest_basis_bil"])
                + _decimal(interest["projected_on_rrp_interest_basis_bil"])
                + max(
                    _decimal(interest["projected_current_remittance_state_bil"]),
                    Decimal("0"),
                )
            )
            demand_converted = _decimal(
                interest["gross_public_interest_current_demand_support_bil"]
            )
            residual_basis = max(private_basis - demand_converted, Decimal("0"))
            household_access = (
                _decimal(assumption.household_safe_asset_stock_share)
                * _decimal(assumption.household_safe_asset_access_conditioner)
            )
            retail_beta = _decimal(assumption.retail_safe_yield_pass_through_beta)
            spend_share = _decimal(assumption.household_safe_yield_current_spend_share)
            household_capture = (
                residual_basis * household_access * retail_beta * spend_share
            )
            mmf_access = (
                max(Decimal("0"), Decimal("1") - household_access)
                * _decimal(assumption.deposit_mmf_substitution_conditioner)
            )
            mmf_offset = residual_basis * mmf_access * retail_beta * spend_share
            credit_drag_basis = (
                gdp
                * _decimal(assumption.contractionary_drag_gdp_share)
                * curve_bp
                / Decimal("100")
                * _decimal(assumption.credit_supply_drag_share)
            )
            mmf_drag = (
                credit_drag_basis
                * mmf_access
                * _decimal(assumption.deposit_mmf_substitution_drag_share)
            )
            paired_net = mmf_offset - mmf_drag
            denominator_overlap = (
                "credit_drag_has_moving_D_overlap_not_added_to_main_n"
                if mmf_drag != 0
                else "no_credit_drag_overlap_at_zero_or_inactive_pair"
            )
            total = firm_cash + household_capture + paired_net
            rows.append(
                {
                    "forecast_residual_numerator_sensitivity_row_id": (
                        "forecast_residual_numerator_sensitivity::"
                        f"{assumption.name}::{year}::{effect['scenario_id']}"
                    ),
                    "fiscal_year": year,
                    "scenario_id": effect["scenario_id"],
                    "baseline_scenario_id": effect["baseline_scenario_id"],
                    "assumption_set": assumption.name,
                    "source_vintage": "cbo_2026_02_fred_latest_cache",
                    "cbo_nominal_gdp_bil": _fmt(gdp),
                    "curve_effective_overlay_bp": _fmt(curve_bp),
                    "selected_moving_denominator_bil": syn[
                        "selected_moving_denominator_bil"
                    ],
                    "selected_delta_denominator_bil": syn[
                        "selected_delta_denominator_bil"
                    ],
                    "firm_liquid_asset_stock_gdp_share": _fmt(firm_stock_share),
                    "firm_liquid_asset_stock_source_status": firm_stock_status,
                    "firm_cash_rate_path_yield_basis_bil": _fmt(firm_cash_basis),
                    "firm_cash_attenuation_share": _fmt(
                        _decimal(assumption.firm_cash_attenuation_share)
                    ),
                    "firm_cash_attenuation_bil": _fmt(firm_cash),
                    "public_interest_private_recipient_cashflow_basis_bil": _fmt(
                        private_basis
                    ),
                    "public_interest_already_demand_converted_bil": _fmt(
                        demand_converted
                    ),
                    "public_interest_residual_cashflow_basis_bil": _fmt(
                        residual_basis
                    ),
                    "household_safe_asset_stock_share": _fmt(
                        _decimal(assumption.household_safe_asset_stock_share)
                    ),
                    "household_safe_asset_access_conditioner": _fmt(
                        _decimal(assumption.household_safe_asset_access_conditioner)
                    ),
                    "retail_safe_yield_pass_through_beta": _fmt(retail_beta),
                    "household_safe_yield_current_spend_share": _fmt(spend_share),
                    "household_safe_yield_capture_bil": _fmt(household_capture),
                    "deposit_mmf_incremental_access_share": _fmt(mmf_access),
                    "deposit_mmf_substitution_conditioner": _fmt(
                        _decimal(assumption.deposit_mmf_substitution_conditioner)
                    ),
                    "deposit_mmf_substitution_offset_bil": _fmt(mmf_offset),
                    "credit_supply_drag_basis_bil": _fmt(credit_drag_basis),
                    "deposit_mmf_substitution_drag_share": _fmt(
                        _decimal(assumption.deposit_mmf_substitution_drag_share)
                    ),
                    "deposit_mmf_substitution_drag_bil": _fmt(mmf_drag),
                    "paired_deposit_mmf_net_sensitivity_bil": _fmt(paired_net),
                    "total_residual_sensitivity_bil": _fmt(total),
                    "baseline_firm_cash_attenuation_bil": "",
                    "delta_firm_cash_attenuation_vs_baseline_bil": "",
                    "baseline_household_safe_yield_capture_bil": "",
                    "delta_household_safe_yield_capture_vs_baseline_bil": "",
                    "baseline_paired_deposit_mmf_net_sensitivity_bil": "",
                    "delta_paired_deposit_mmf_net_sensitivity_vs_baseline_bil": "",
                    "baseline_total_residual_sensitivity_bil": "",
                    "delta_total_residual_sensitivity_vs_baseline_bil": "",
                    "firm_cash_status": (
                        "bounded_sensitivity_source_backed_stock_context"
                    ),
                    "safe_yield_status": (
                        "bounded_residual_sensitivity_zero_if_assumption_inactive"
                    ),
                    "deposit_mmf_pairing_status": (
                        "paired_offset_and_drag_never_unpaired_offset"
                    ),
                    "denominator_overlap_status": denominator_overlap,
                    "composition_rule": (
                        "residual_basis_after_public_interest_demand_conversion;"
                        "safe_asset_offset_obsolete_not_added;"
                        "firm_cash_cushion_replacement_only"
                    ),
                    "allowed_use": "forecast_residual_numerator_sensitivity",
                    "blocked_use": (
                        "canonical_headline_promotion;unpaired_mmf_offset;"
                        "same_cashflow_dollar_converted_twice;"
                        "credit_drag_double_count_against_moving_D"
                    ),
                    "canonical_ratio_entry": "false",
                    "enters_main_ratio": "false",
                    "evidence_mode_enabled": "false",
                }
            )
    baseline_by_key = {
        (row["assumption_set"], row["fiscal_year"]): row
        for row in rows
        if row["scenario_id"] == row["baseline_scenario_id"]
    }
    for row in rows:
        baseline = baseline_by_key.get((row["assumption_set"], row["fiscal_year"]))
        if baseline is None:
            raise ForecastModelReadoutError(
                "missing residual baseline row for "
                f"{row['assumption_set']}::{row['fiscal_year']}"
            )
        _add_residual_delta_fields(row, baseline)
    return sorted(
        rows,
        key=lambda row: (
            row["assumption_set"],
            int(row["fiscal_year"]),
            row["scenario_id"],
        ),
    ), source_list


def forecast_composition_surface_rows(
    *,
    timed_beta_rows: Iterable[Mapping[str, str]],
    public_interest_rows: Iterable[Mapping[str, str]],
    residual_sensitivity_rows: Iterable[Mapping[str, str]],
) -> list[dict[str, str]]:
    """Compose candidate forecast numerators against the selected moving D."""

    timed = _by_key_with_optional_axis(
        timed_beta_rows,
        "timed beta",
        axis_field="beta_path_id",
    )
    public_interest = _by_key(public_interest_rows, "public-interest net block")
    residual_by_key: dict[
        tuple[str, str, str],
        dict[str, str],
    ] = {
        (row["assumption_set"], row["scenario_id"], row["fiscal_year"]): dict(row)
        for row in residual_sensitivity_rows
    }
    residual_assumption_sets = sorted(
        {row["assumption_set"] for row in residual_sensitivity_rows}
    )
    rows: list[dict[str, str]] = []
    for (_scenario_id, _year, _beta_path_id), timed_row in timed.items():
        key = (timed_row["scenario_id"], timed_row["fiscal_year"])
        public = _required(public_interest, key, "public-interest net block")
        rows.append(
            _composition_surface_row(
                timed_row,
                public_interest=public,
                composition_case_id="first_forecast_current",
                composition_case_label="First forecast numerator",
                residual_assumption_set="",
                residual_sensitivity_delta=Decimal("0"),
                use_public_interest_replacement=False,
            )
        )
        rows.append(
            _composition_surface_row(
                timed_row,
                public_interest=public,
                composition_case_id="public_interest_replacement",
                composition_case_label="Public-interest replacement numerator",
                residual_assumption_set="",
                residual_sensitivity_delta=Decimal("0"),
                use_public_interest_replacement=True,
            )
        )
        for assumption_set in residual_assumption_sets:
            residual = residual_by_key.get(
                (assumption_set, timed_row["scenario_id"], timed_row["fiscal_year"])
            )
            if residual is None:
                continue
            rows.append(
                _composition_surface_row(
                    timed_row,
                    public_interest=public,
                    composition_case_id=(
                        "public_interest_plus_residual_delta::"
                        f"{assumption_set}"
                    ),
                    composition_case_label=(
                        "Public-interest replacement plus residual sensitivity "
                        "delta"
                    ),
                    residual_assumption_set=assumption_set,
                    residual_sensitivity_delta=_decimal(
                        residual[
                            "delta_total_residual_sensitivity_vs_baseline_bil"
                        ]
                    ),
                    use_public_interest_replacement=True,
                )
            )
    baseline_by_key = {
        (row["fiscal_year"], row["beta_path_id"], row["composition_case_id"]): row
        for row in rows
        if row["scenario_id"] == row["baseline_scenario_id"]
    }
    for row in rows:
        baseline = baseline_by_key.get(
            (row["fiscal_year"], row["beta_path_id"], row["composition_case_id"])
        )
        if baseline is None:
            raise ForecastModelReadoutError(
                "missing composition baseline for "
                f"{row['fiscal_year']}::{row['beta_path_id']}::"
                f"{row['composition_case_id']}"
            )
        _add_composition_delta_fields(row, baseline)
    return sorted(
        rows,
        key=lambda row: (
            row["composition_case_id"],
            row["beta_path_id"],
            int(row["fiscal_year"]),
            row["scenario_id"],
        ),
    )


def central_forecast_surface_rows(
    composition_surface_rows: Iterable[Mapping[str, str]],
) -> list[dict[str, str]]:
    """Select the central forecast and attach sensitivity comparisons."""

    rows = list(composition_surface_rows)
    by_key = {
        (
            row["composition_case_id"],
            row["beta_path_id"],
            row["scenario_id"],
            row["fiscal_year"],
        ): dict(row)
        for row in rows
    }
    central_rows = [
        row
        for row in rows
        if row["composition_case_id"] == "public_interest_replacement"
        and row["beta_path_id"] == "normal_forward_constant"
    ]
    out: list[dict[str, str]] = []
    for central in central_rows:
        scenario_id = central["scenario_id"]
        year = central["fiscal_year"]
        central_ratio = _decimal(central["composition_ratewall_ratio"])
        first = _required_composition_case(
            by_key,
            case_id="first_forecast_current",
            beta_path_id="normal_forward_constant",
            scenario_id=scenario_id,
            fiscal_year=year,
        )
        residual_base = _required_composition_case(
            by_key,
            case_id="public_interest_plus_residual_delta::literature_calibrated_base",
            beta_path_id="normal_forward_constant",
            scenario_id=scenario_id,
            fiscal_year=year,
        )
        residual_paired = _required_composition_case(
            by_key,
            case_id=(
                "public_interest_plus_residual_delta::"
                "assumption_mode_deposit_mmf_paired_entry"
            ),
            beta_path_id="normal_forward_constant",
            scenario_id=scenario_id,
            fiscal_year=year,
        )
        latest = _required_composition_case(
            by_key,
            case_id="public_interest_replacement",
            beta_path_id="latest_rolling_from_fy2032",
            scenario_id=scenario_id,
            fiscal_year=year,
        )
        pooled = _required_composition_case(
            by_key,
            case_id="public_interest_replacement",
            beta_path_id="pooled_full_from_fy2032_review",
            scenario_id=scenario_id,
            fiscal_year=year,
        )
        out.append(
            {
                "central_forecast_surface_row_id": (
                    "central_forecast_surface::"
                    f"{year}::{scenario_id}"
                ),
                "fiscal_year": year,
                "scenario_id": scenario_id,
                "baseline_scenario_id": central["baseline_scenario_id"],
                "central_beta_path_id": "normal_forward_constant",
                "central_composition_case_id": "public_interest_replacement",
                "central_n_bil": central["composition_n_bil"],
                "central_moving_denominator_bil": central[
                    "selected_moving_denominator_bil"
                ],
                "central_ratewall_ratio": central["composition_ratewall_ratio"],
                "baseline_central_n_bil": central["baseline_composition_n_bil"],
                "baseline_central_moving_denominator_bil": central[
                    "baseline_selected_moving_denominator_bil"
                ],
                "baseline_central_ratewall_ratio": central[
                    "baseline_composition_ratewall_ratio"
                ],
                "delta_central_n_vs_baseline_bil": central[
                    "delta_composition_n_vs_baseline_bil"
                ],
                "delta_central_moving_denominator_vs_baseline_bil": central[
                    "selected_delta_denominator_bil"
                ],
                "delta_central_ratewall_ratio_vs_baseline": central[
                    "delta_composition_ratewall_ratio_vs_baseline"
                ],
                "wall_hit_under_central_forecast": central[
                    "wall_hit_under_composition"
                ],
                "first_forecast_ratewall_ratio": first["composition_ratewall_ratio"],
                "delta_first_forecast_ratewall_ratio_vs_central": _fmt(
                    _decimal(first["composition_ratewall_ratio"]) - central_ratio
                ),
                "residual_base_ratewall_ratio": residual_base[
                    "composition_ratewall_ratio"
                ],
                "delta_residual_base_ratewall_ratio_vs_central": _fmt(
                    _decimal(residual_base["composition_ratewall_ratio"])
                    - central_ratio
                ),
                "residual_paired_ratewall_ratio": residual_paired[
                    "composition_ratewall_ratio"
                ],
                "delta_residual_paired_ratewall_ratio_vs_central": _fmt(
                    _decimal(residual_paired["composition_ratewall_ratio"])
                    - central_ratio
                ),
                "latest_rolling_beta_ratewall_ratio": latest[
                    "composition_ratewall_ratio"
                ],
                "delta_latest_rolling_beta_ratewall_ratio_vs_central": _fmt(
                    _decimal(latest["composition_ratewall_ratio"]) - central_ratio
                ),
                "pooled_full_beta_ratewall_ratio": pooled[
                    "composition_ratewall_ratio"
                ],
                "delta_pooled_full_beta_ratewall_ratio_vs_central": _fmt(
                    _decimal(pooled["composition_ratewall_ratio"]) - central_ratio
                ),
                "central_choice_status": (
                    "selected_model_surface_public_interest_replacement_"
                    "normal_forward_beta"
                ),
                "sensitivity_rule": (
                    "first_forecast_residual_and_timed_beta_paths_are_"
                    "sensitivity_surfaces_not_central"
                ),
                "allowed_use": "central_forecast_model_scenario_surface",
                "blocked_use": (
                    "canonical_headline_promotion;release_headline_claim;"
                    "evidence_mode_claim;denominator_prior_update;beta_prior_update;"
                    "chi_prior_update"
                ),
                "canonical_ratio_entry": "false",
                "enters_main_ratio": "false",
                "evidence_mode_enabled": "false",
            }
        )
    return sorted(out, key=lambda row: (int(row["fiscal_year"]), row["scenario_id"]))


def central_scenario_interpretation_rows(
    central_forecast_rows: Iterable[Mapping[str, str]],
) -> list[dict[str, str]]:
    """Add plain-language mechanism and sensitivity diagnostics."""

    rows = list(central_forecast_rows)
    out: list[dict[str, str]] = []
    for row in rows:
        baseline_n = _decimal(row["baseline_central_n_bil"])
        baseline_d = _decimal(row["baseline_central_moving_denominator_bil"])
        current_d = _decimal(row["central_moving_denominator_bil"])
        delta_n = _decimal(row["delta_central_n_vs_baseline_bil"])
        delta_d = _decimal(row["delta_central_moving_denominator_vs_baseline_bil"])
        baseline_ratio = _decimal(row["baseline_central_ratewall_ratio"])
        numerator_only = (baseline_n + delta_n) / baseline_d - baseline_ratio
        denominator_only = baseline_n / current_d - baseline_ratio
        primary_driver = _primary_driver(numerator_only, denominator_only)
        sensitivity_values = {
            "central": _decimal(row["central_ratewall_ratio"]),
            "first_forecast": _decimal(row["first_forecast_ratewall_ratio"]),
            "residual_base": _decimal(row["residual_base_ratewall_ratio"]),
            "residual_paired": _decimal(row["residual_paired_ratewall_ratio"]),
            "latest_rolling_beta": _decimal(row["latest_rolling_beta_ratewall_ratio"]),
            "pooled_full_beta": _decimal(row["pooled_full_beta_ratewall_ratio"]),
        }
        central_ratio = sensitivity_values["central"]
        largest_case, largest_delta = max(
            (
                (case, value - central_ratio)
                for case, value in sensitivity_values.items()
                if case != "central"
            ),
            key=lambda item: abs(item[1]),
        )
        minimum = min(sensitivity_values.values())
        maximum = max(sensitivity_values.values())
        out.append(
            {
                "central_scenario_interpretation_row_id": (
                    "central_scenario_interpretation::"
                    f"{row['fiscal_year']}::{row['scenario_id']}"
                ),
                "fiscal_year": row["fiscal_year"],
                "scenario_id": row["scenario_id"],
                "baseline_scenario_id": row["baseline_scenario_id"],
                "central_ratewall_ratio": row["central_ratewall_ratio"],
                "delta_central_ratewall_ratio_vs_baseline": row[
                    "delta_central_ratewall_ratio_vs_baseline"
                ],
                "delta_central_n_vs_baseline_bil": row[
                    "delta_central_n_vs_baseline_bil"
                ],
                "delta_central_moving_denominator_vs_baseline_bil": row[
                    "delta_central_moving_denominator_vs_baseline_bil"
                ],
                "numerator_only_delta_ratewall_ratio": _fmt(numerator_only),
                "denominator_only_delta_ratewall_ratio": _fmt(denominator_only),
                "primary_driver": primary_driver,
                "scenario_direction": _scenario_direction(
                    _decimal(row["delta_central_ratewall_ratio_vs_baseline"])
                ),
                "mechanism_summary": _mechanism_summary(
                    delta_n=delta_n,
                    delta_d=delta_d,
                    numerator_only=numerator_only,
                    denominator_only=denominator_only,
                    primary_driver=primary_driver,
                ),
                "sensitivity_min_ratewall_ratio": _fmt(minimum),
                "sensitivity_max_ratewall_ratio": _fmt(maximum),
                "sensitivity_width_ratewall_ratio": _fmt(maximum - minimum),
                "largest_sensitivity_case": largest_case,
                "largest_sensitivity_delta_vs_central": _fmt(largest_delta),
                "allowed_use": "central_forecast_plain_model_interpretation",
                "blocked_use": (
                    "canonical_headline_promotion;release_headline_claim;"
                    "evidence_mode_claim;statistical_significance_claim"
                ),
                "canonical_ratio_entry": "false",
                "enters_main_ratio": "false",
                "evidence_mode_enabled": "false",
            }
        )
    return sorted(out, key=lambda item: (int(item["fiscal_year"]), item["scenario_id"]))


def forecast_scenario_sufficiency_rows(
    *,
    effect_rows: Iterable[Mapping[str, str]],
    scenario_config_rows: Iterable[Mapping[str, str]],
    central_interpretation_rows: Iterable[Mapping[str, str]],
) -> list[dict[str, str]]:
    """Classify current scenario coverage and high-value remaining gaps."""

    effects = list(effect_rows)
    configs = list(scenario_config_rows)
    run_ids = {row["scenario_id"] for row in effects}
    config_by_id = {row["scenario_id"]: row for row in configs}
    central_2036 = {
        row["scenario_id"]: row
        for row in central_interpretation_rows
        if row["fiscal_year"] == "2036"
    }
    scenario_ids = sorted(run_ids | set(config_by_id))
    rows = [
        _scenario_sufficiency_row(
            scenario_id=scenario_id,
            config=config_by_id.get(scenario_id, {}),
            run_in_suite=scenario_id in run_ids,
            central_2036=central_2036.get(scenario_id),
        )
        for scenario_id in scenario_ids
    ]
    return sorted(rows, key=lambda row: (_scenario_sufficiency_sort(row), row["scenario_id"]))


def residual_sensitivity_source_rows(
    source_cache_dir: str | Path | None = DEFAULT_FED_SOURCE_CACHE_DIR,
    *,
    as_of_date: date | None = None,
) -> list[dict[str, str]]:
    """Summarize cached FRED rows used by residual numerator sensitivities."""

    if source_cache_dir is None:
        return []
    return _forecast_source_rows(
        source_cache_dir=source_cache_dir,
        series=RESIDUAL_SENSITIVITY_FRED_SERIES,
        row_id_field="forecast_residual_sensitivity_source_row_id",
        row_id_prefix="forecast_residual_sensitivity_source",
        as_of_date=as_of_date,
        latest_average_window_size=4,
    )


def fed_forecast_source_rows(
    source_cache_dir: str | Path | None = DEFAULT_FED_SOURCE_CACHE_DIR,
    *,
    as_of_date: date | None = None,
) -> list[dict[str, str]]:
    """Summarize cached FRED source CSVs into compact forecast source rows."""

    return _forecast_source_rows(
        source_cache_dir=source_cache_dir,
        series=FRED_SERIES,
        row_id_field="forecast_fed_liability_source_row_id",
        row_id_prefix="forecast_fed_liability_source",
        as_of_date=as_of_date,
        latest_average_window_size=13,
    )


def _forecast_source_rows(
    *,
    source_cache_dir: str | Path | None,
    series: Mapping[str, Mapping[str, str]],
    row_id_field: str,
    row_id_prefix: str,
    as_of_date: date | None,
    latest_average_window_size: int,
) -> list[dict[str, str]]:
    """Summarize cached FRED source CSVs into compact source rows."""

    if source_cache_dir is None:
        return []
    cutoff = as_of_date or date.today()
    source_dir = Path(source_cache_dir)
    rows: list[dict[str, str]] = []
    for series_id, meta in series.items():
        path = source_dir / f"{series_id}.csv"
        if not path.exists():
            continue
        observations = [
            item for item in _read_fred_csv(path) if item[0] <= cutoff
        ]
        if not observations:
            continue
        window = observations[-latest_average_window_size:]
        latest_date, latest_value = observations[-1]
        avg = sum((value for _date, value in window), Decimal("0")) / Decimal(
            len(window)
        )
        rows.append(
            {
                row_id_field: f"{row_id_prefix}::{series_id}",
                "series_id": series_id,
                "series_label": meta["label"],
                "source_url": FRED_GRAPH_URL.format(series_id=series_id),
                "source_cache_path": str(path),
                "observation_count": str(len(observations)),
                "latest_observation_date": latest_date.isoformat(),
                "latest_observation_value": _fmt(latest_value),
                "latest_average_window_observation_count": str(len(window)),
                "latest_average_value": _fmt(avg),
                "unit": meta["unit"],
                "projection_use": meta["projection_use"],
                "source_status": "pass_cached_official_fred_csv",
            }
        )
    return rows


def cbo_fiscal_macro_rows(
    workbook_path: str | Path = CBO_ECONOMIC_PROJECTIONS_XLSX,
) -> list[dict[str, str]]:
    """Read CBO fiscal-year nominal GDP and short-rate path from the workbook."""

    path = Path(workbook_path)
    if not path.exists():
        raise ForecastModelReadoutError(f"missing CBO economic workbook: {path}")
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["3. Fiscal Year"]
        year_values = _cbo_sheet_year_values(sheet)
        gdp_values = _cbo_sheet_values(sheet, "Gross domestic product (GDP)")
        short_rate_values = _cbo_sheet_values(sheet, "Federal funds rate")
    finally:
        workbook.close()
    rows = []
    for year_value, gdp, short_rate in zip(
        year_values,
        gdp_values,
        short_rate_values,
        strict=False,
    ):
        year = _year_from_value(year_value)
        if year is None:
            continue
        if 2027 <= year <= 2036:
            rows.append(
                {
                    "fiscal_year": str(year),
                    "cbo_nominal_gdp_bil": _fmt(_decimal(gdp)),
                    "cbo_short_rate_pct": _fmt(_decimal(short_rate)),
                    "source_status": "pass_cbo_2026_02_fiscal_year_workbook",
                }
            )
    if len(rows) != 10:
        raise ForecastModelReadoutError(
            f"expected 10 CBO macro rows for FY2027-FY2036, got {len(rows)}"
        )
    return rows


def cbo_remittance_projection_rows(
    workbook_path: str | Path = CBO_BUDGET_PROJECTIONS_XLSX,
) -> list[dict[str, str]]:
    """Return CBO remittance rows only when a level forecast is clearly available."""

    path = Path(workbook_path)
    if not path.exists():
        return []
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        # The current workbook has a Table 5-1 "changes since baseline" row for
        # Federal Reserve remittances, not a clean level forecast. Keep the final
        # block fail-closed to zero unless a level row appears in a future source.
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                label = str(row[0]).strip() if row and row[0] is not None else ""
                if label.lower() == "federal reserve remittances level":
                    years = [int(value) for value in row[1:11]]
                    values = [_decimal(value) for value in row[11:21]]
                    return [
                        {
                            "fiscal_year": str(year),
                            "cbo_federal_reserve_remittance_bil": _fmt(value),
                        }
                        for year, value in zip(years, values, strict=True)
                    ]
    finally:
        workbook.close()
    return []


def timed_beta_path_rows(
    *,
    effect_rows: Iterable[Mapping[str, str]],
    summary_rows: Iterable[Mapping[str, str]],
    synthesis_rows: Iterable[Mapping[str, str]],
    materiality_rows: Iterable[Mapping[str, str]],
) -> list[dict[str, str]]:
    """Recompute the active scenario surface under a small set of beta paths."""

    effects = _by_key(effect_rows, "scenario effect")
    summaries = _by_key(summary_rows, "model summary")
    synthesis = _by_key(synthesis_rows, "model synthesis")
    materiality = _by_key(materiality_rows, "materiality", require_unique=False)
    normal_rows_by_key: dict[tuple[str, str], dict[str, str]] = {}
    out: list[dict[str, str]] = []
    for beta_path in TIMED_BETA_PATHS:
        for key, summary in summaries.items():
            fiscal_year = int(summary["fiscal_year"])
            beta_label, beta, beta_source = _beta_for_path(beta_path, fiscal_year)
            syn = _required(synthesis, key, "model synthesis")
            effect = _required(effects, key, "scenario effect")
            baseline_key = (summary["baseline_scenario_id"], summary["fiscal_year"])
            baseline_effect = _required(effects, baseline_key, "baseline effect")
            baseline_syn = _required(synthesis, baseline_key, "baseline synthesis")
            mat = materiality.get(key, {})
            row = _timed_beta_row(
                summary,
                effect=effect,
                baseline_effect=baseline_effect,
                synthesis=syn,
                baseline_synthesis=baseline_syn,
                materiality=mat,
                beta_path=beta_path,
                beta_label=beta_label,
                beta=beta,
                beta_source_status=beta_source,
            )
            if beta_path["beta_path_id"] == "normal_forward_constant":
                normal_rows_by_key[key] = row
            else:
                normal = _required(
                    normal_rows_by_key,
                    key,
                    "normal-forward timed beta row",
                )
                row["delta_ratewall_ratio_vs_normal_forward_path"] = _fmt(
                    _decimal(row["level_ratewall_ratio_recomputed"])
                    - _decimal(normal["level_ratewall_ratio_recomputed"])
                )
            out.append(row)
    return sorted(
        out,
        key=lambda row: (
            row["beta_path_id"],
            int(row["fiscal_year"]),
            row["scenario_id"],
        ),
    )


def forecast_channel_classification_rows() -> list[dict[str, str]]:
    """Classify numerator channels so deferred current/historical pieces are visible."""

    rows: list[dict[str, str]] = []
    for channel_id, label, status, reason in ACTIVE_FORECAST_CHANNELS:
        roles = _forecast_channel_entry_roles(channel_id, "included", {})
        rows.append(
            {
                "forecast_channel_classification_row_id": (
                    f"forecast_channel_classification::{channel_id}"
                ),
                "channel_id": channel_id,
                "channel_label": label,
                "current_or_historical_channel_status": "active_ratewall_channel",
                "first_10y_forecast_status": status,
                "first_forecast_entry_role": roles["first_forecast_entry_role"],
                "selected_central_entry_role": roles["selected_central_entry_role"],
                "public_interest_block_role": roles["public_interest_block_role"],
                "classification": "included",
                "reason": reason,
                "next_model_requirement": "keep_identity_tests_on_first_forecast_numerator",
                "allowed_use": "forecast_scope_control",
                "blocked_use": "claim_that_deferred_channels_are_zero_or_obsolete",
            }
        )
    for channel_id in DEFERRED_FORECAST_CHANNELS:
        treatment = DEFERRED_CHANNEL_FINAL_TREATMENT[channel_id]
        roles = _forecast_channel_entry_roles(
            channel_id,
            treatment["classification"],
            treatment,
        )
        rows.append(
            {
                "forecast_channel_classification_row_id": (
                    f"forecast_channel_classification::{channel_id}"
                ),
                "channel_id": channel_id,
                "channel_label": channel_id.replace("_", " "),
                "current_or_historical_channel_status": (
                    "tracked_in_current_or_historical_ratewall_numerator"
                ),
                "first_10y_forecast_status": "not_in_first_10y_tdcsim_cbo_forecast",
                "first_forecast_entry_role": roles["first_forecast_entry_role"],
                "selected_central_entry_role": roles["selected_central_entry_role"],
                "public_interest_block_role": roles["public_interest_block_role"],
                "classification": treatment["classification"],
                "reason": treatment["reason"],
                "next_model_requirement": treatment["next_model_requirement"],
                "allowed_use": "forecast_scope_control",
                "blocked_use": treatment["blocked_use"],
            }
        )
    return rows


def _forecast_channel_entry_roles(
    channel_id: str,
    classification: str,
    treatment: Mapping[str, str],
) -> dict[str, str]:
    """Separate first-forecast rows from selected central replacement-block roles."""

    if channel_id == "tdc_ex_overlap_current_demand_support":
        return {
            "first_forecast_entry_role": "standalone_final_n_term",
            "selected_central_entry_role": "standalone_final_n_term",
            "public_interest_block_role": "not_applicable",
        }
    if channel_id in {
        "direct_treasury_interest_support",
        "bank_treasury_interest_support",
    }:
        return {
            "first_forecast_entry_role": "standalone_final_n_term",
            "selected_central_entry_role": "replacement_block_input_not_standalone",
            "public_interest_block_role": "legacy_interest_input",
        }
    if classification == "included_as_public_interest_replacement_block":
        return {
            "first_forecast_entry_role": "not_in_first_forecast",
            "selected_central_entry_role": "standalone_final_n_term",
            "public_interest_block_role": treatment.get(
                "public_interest_block_role",
                "block_output_standalone_final_n_term",
            ),
        }
    if classification.startswith("integrated_in_public_interest_net_block"):
        return {
            "first_forecast_entry_role": "not_in_first_forecast",
            "selected_central_entry_role": "replacement_block_input_not_standalone",
            "public_interest_block_role": treatment.get(
                "public_interest_block_role",
                "public_interest_block_input",
            ),
        }
    if classification == "replaced_by_current_tdcsim_holder_filter":
        return {
            "first_forecast_entry_role": "not_in_first_forecast",
            "selected_central_entry_role": "diagnostic_or_guard_not_standalone",
            "public_interest_block_role": "holder_filter_guard",
        }
    if classification in {
        "projection_required_as_bounded_sensitivity",
        "projection_required_as_bounded_residual_sensitivity",
        "projection_required_as_paired_bounded_sensitivity",
    }:
        return {
            "first_forecast_entry_role": "not_in_first_forecast",
            "selected_central_entry_role": "sensitivity_only_not_central",
            "public_interest_block_role": "not_applicable",
        }
    return {
        "first_forecast_entry_role": "not_in_first_forecast",
        "selected_central_entry_role": "not_selected_central_n_term",
        "public_interest_block_role": treatment.get(
            "public_interest_block_role",
            "not_applicable",
        ),
    }


def forecast_numerator_channel_plan_rows(
    channel_rows: Sequence[Mapping[str, str]] = (),
) -> list[dict[str, str]]:
    """Return the final-version plan for unresolved deferred numerator channels."""

    if channel_rows:
        classifications = {row["channel_id"]: row["classification"] for row in channel_rows}
        labels = {row["channel_id"]: row["channel_label"] for row in channel_rows}
    else:
        classifications = {
            channel_id: DEFERRED_CHANNEL_FINAL_TREATMENT[channel_id]["classification"]
            for channel_id in REMAINING_NUMERATOR_CHANNEL_PLAN
        }
        labels = {
            channel_id: channel_id.replace("_", " ")
            for channel_id in REMAINING_NUMERATOR_CHANNEL_PLAN
        }
    rows: list[dict[str, str]] = []
    for channel_id, plan in REMAINING_NUMERATOR_CHANNEL_PLAN.items():
        rows.append(
            {
                "forecast_numerator_channel_plan_row_id": (
                    f"forecast_numerator_channel_plan::{channel_id}"
                ),
                "channel_id": channel_id,
                "channel_label": labels.get(channel_id, channel_id.replace("_", " ")),
                "current_classification": classifications.get(
                    channel_id,
                    DEFERRED_CHANNEL_FINAL_TREATMENT[channel_id]["classification"],
                ),
                "final_central_status": plan["final_central_status"],
                "materiality_tier": plan["materiality_tier"],
                "forecast_route": plan["forecast_route"],
                "assumption_basis": plan["assumption_basis"],
                "calibration_need": plan["calibration_need"],
                "double_count_guard": plan["double_count_guard"],
                "admission_test": plan["admission_test"],
                "next_model_action": plan["next_model_action"],
                "allowed_use": "final_numerator_gap_closure_plan",
                "blocked_use": plan["blocked_use"],
            }
        )
    return rows


def zero_low_apr_credit_materiality_rows() -> list[dict[str, str]]:
    """Screen zero/low-APR credit products without admitting them to central N."""

    rows: list[dict[str, str]] = []
    for source in ZERO_LOW_APR_CREDIT_SCREEN_ROWS:
        stock = _optional_decimal(source["outstanding_stock_bil"])
        rate_wedge = _optional_decimal(source["candidate_rate_wedge_pct"])
        relief = stock * rate_wedge / Decimal("100") if stock is not None and rate_wedge is not None else None
        product_segment = source["product_segment"]
        rows.append(
            {
                "zero_low_apr_credit_materiality_row_id": (
                    f"zero_low_apr_credit_materiality::{product_segment}"
                ),
                "product_segment": product_segment,
                "source_vintage": source["source_vintage"],
                "source_url": source["source_url"],
                "source_metric": source["source_metric"],
                "annual_originations_bil": source["annual_originations_bil"],
                "outstanding_stock_bil": source["outstanding_stock_bil"],
                "zero_low_apr_share": source["zero_low_apr_share"],
                "duration_months": source["duration_months"],
                "candidate_rate_wedge_pct": source["candidate_rate_wedge_pct"],
                "screen_relief_bil": _fmt(relief) if relief is not None else "",
                "screen_status": source["screen_status"],
                "central_n_treatment": source["central_n_treatment"],
                "materiality_decision": source["materiality_decision"],
                "next_model_action": source["next_model_action"],
                "allowed_use": "zero_low_apr_credit_materiality_screen_only",
                "blocked_use": source["blocked_use"],
            }
        )
    return rows


def write_forecast_model_readout_outputs(
    output_dir: str | Path,
    *,
    timed_beta_rows: Sequence[Mapping[str, str]],
    channel_rows: Sequence[Mapping[str, str]],
    numerator_channel_plan_rows: Sequence[Mapping[str, str]] = (),
    zero_low_apr_credit_rows: Sequence[Mapping[str, str]] = (),
    public_interest_rows: Sequence[Mapping[str, str]] = (),
    fed_source_rows: Sequence[Mapping[str, str]] = (),
    residual_sensitivity_rows: Sequence[Mapping[str, str]] = (),
    residual_source_rows: Sequence[Mapping[str, str]] = (),
    composition_surface_rows: Sequence[Mapping[str, str]] = (),
    central_forecast_rows: Sequence[Mapping[str, str]] = (),
    central_interpretation_rows: Sequence[Mapping[str, str]] = (),
    scenario_sufficiency_rows: Sequence[Mapping[str, str]] = (),
) -> dict[str, Path]:
    """Write ten-year forecast CSV, PNG, and Markdown readout artifacts."""

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    paths = {
        "timed_beta_csv": out / "ratewall_forecast_timed_beta_paths.csv",
        "public_interest_net_block_csv": (
            out / "ratewall_forecast_public_interest_net_block.csv"
        ),
        "fed_liability_sources_csv": (
            out / "ratewall_forecast_fed_liability_sources.csv"
        ),
        "residual_numerator_sensitivity_csv": (
            out / "ratewall_forecast_residual_numerator_sensitivity.csv"
        ),
        "residual_sensitivity_sources_csv": (
            out / "ratewall_forecast_residual_sensitivity_sources.csv"
        ),
        "composition_surface_csv": (
            out / "ratewall_forecast_composition_surface.csv"
        ),
        "central_forecast_surface_csv": (
            out / "ratewall_forecast_central_scenario_surface.csv"
        ),
        "central_interpretation_csv": (
            out / "ratewall_forecast_central_scenario_interpretation.csv"
        ),
        "scenario_sufficiency_csv": (
            out / "ratewall_forecast_scenario_sufficiency_review.csv"
        ),
        "channel_classification_csv": (
            out / "ratewall_forecast_channel_classification.csv"
        ),
        "numerator_channel_plan_csv": (
            out / "ratewall_forecast_numerator_channel_plan.csv"
        ),
        "zero_low_apr_credit_materiality_csv": (
            out / "ratewall_forecast_zero_low_apr_credit_materiality.csv"
        ),
        "readout_md": out / "ratewall_forecast_10y_readout.md",
        "png_ratewall_paths": out / "forecast_01_ratewall_paths.png",
        "png_timed_beta_effect": out / "forecast_02_timed_beta_effect.png",
        "png_components": out / "forecast_03_components_2036.png",
        "png_channel_scope": out / "forecast_04_channel_scope.png",
        "png_composition_surface": out / "forecast_05_composition_surface_2036.png",
        "png_central_surface": out / "forecast_06_central_surface_2036.png",
        "png_central_baseline_path": out / "forecast_07_central_baseline_path.png",
        "png_central_sensitivity_spread": (
            out / "forecast_08_central_sensitivity_spread_2036.png"
        ),
        "png_scenario_sufficiency": out / "forecast_09_scenario_sufficiency.png",
    }
    _write_csv(paths["timed_beta_csv"], TIMED_BETA_PATH_FIELDS, timed_beta_rows)
    _write_csv(
        paths["public_interest_net_block_csv"],
        PUBLIC_INTEREST_NET_BLOCK_FIELDS,
        public_interest_rows,
    )
    _write_csv(
        paths["fed_liability_sources_csv"],
        FED_FORECAST_SOURCE_FIELDS,
        fed_source_rows,
    )
    _write_csv(
        paths["residual_numerator_sensitivity_csv"],
        RESIDUAL_NUMERATOR_SENSITIVITY_FIELDS,
        residual_sensitivity_rows,
    )
    _write_csv(
        paths["residual_sensitivity_sources_csv"],
        RESIDUAL_SENSITIVITY_SOURCE_FIELDS,
        residual_source_rows,
    )
    _write_csv(
        paths["composition_surface_csv"],
        FORECAST_COMPOSITION_SURFACE_FIELDS,
        composition_surface_rows,
    )
    _write_csv(
        paths["central_forecast_surface_csv"],
        CENTRAL_FORECAST_SURFACE_FIELDS,
        central_forecast_rows,
    )
    _write_csv(
        paths["central_interpretation_csv"],
        CENTRAL_SCENARIO_INTERPRETATION_FIELDS,
        central_interpretation_rows,
    )
    _write_csv(
        paths["scenario_sufficiency_csv"],
        FORECAST_SCENARIO_SUFFICIENCY_FIELDS,
        scenario_sufficiency_rows,
    )
    _write_csv(
        paths["channel_classification_csv"],
        FORECAST_CHANNEL_CLASSIFICATION_FIELDS,
        channel_rows,
    )
    _write_csv(
        paths["numerator_channel_plan_csv"],
        FORECAST_NUMERATOR_CHANNEL_PLAN_FIELDS,
        numerator_channel_plan_rows,
    )
    _write_csv(
        paths["zero_low_apr_credit_materiality_csv"],
        ZERO_LOW_APR_CREDIT_MATERIALITY_FIELDS,
        zero_low_apr_credit_rows,
    )
    paths["readout_md"].write_text(
        forecast_model_readout_markdown(
            timed_beta_rows,
            channel_rows,
            numerator_channel_plan_rows=numerator_channel_plan_rows,
            zero_low_apr_credit_rows=zero_low_apr_credit_rows,
            public_interest_rows=public_interest_rows,
            residual_sensitivity_rows=residual_sensitivity_rows,
            composition_surface_rows=composition_surface_rows,
            central_forecast_rows=central_forecast_rows,
            central_interpretation_rows=central_interpretation_rows,
            scenario_sufficiency_rows=scenario_sufficiency_rows,
        ),
        encoding="utf-8",
    )
    _write_pngs(
        paths,
        timed_beta_rows,
        channel_rows,
        composition_surface_rows,
        central_forecast_rows,
        central_interpretation_rows,
        scenario_sufficiency_rows,
    )
    return paths


def forecast_model_readout_markdown(
    timed_beta_rows: Sequence[Mapping[str, str]],
    channel_rows: Sequence[Mapping[str, str]],
    *,
    numerator_channel_plan_rows: Sequence[Mapping[str, str]] = (),
    zero_low_apr_credit_rows: Sequence[Mapping[str, str]] = (),
    public_interest_rows: Sequence[Mapping[str, str]] = (),
    residual_sensitivity_rows: Sequence[Mapping[str, str]] = (),
    composition_surface_rows: Sequence[Mapping[str, str]] = (),
    central_forecast_rows: Sequence[Mapping[str, str]] = (),
    central_interpretation_rows: Sequence[Mapping[str, str]] = (),
    scenario_sufficiency_rows: Sequence[Mapping[str, str]] = (),
) -> str:
    """Return a concise economist-facing ten-year forecast readout."""

    normal = [
        row
        for row in timed_beta_rows
        if row["beta_path_id"] == "normal_forward_constant"
    ]
    years = sorted({int(row["fiscal_year"]) for row in normal})
    if not years:
        raise ForecastModelReadoutError("timed beta rows are empty")
    first_year = str(years[0])
    last_year = str(years[-1])
    baseline_first = _row_for(normal, "cbo_baseline_noop_v1", first_year)
    baseline_last = _row_for(normal, "cbo_baseline_noop_v1", last_year)
    central_last_rows = [
        row for row in central_forecast_rows if row["fiscal_year"] == last_year
    ]
    central_baseline_last = (
        _row_for(central_last_rows, "cbo_baseline_noop_v1", last_year)
        if central_last_rows
        else None
    )
    largest_central = sorted(
        [
            row
            for row in central_last_rows
            if row["scenario_id"] != row["baseline_scenario_id"]
        ],
        key=lambda row: abs(
            _decimal(row["delta_central_ratewall_ratio_vs_baseline"])
        ),
        reverse=True,
    )[:5]
    central_interpretation_last = [
        row
        for row in central_interpretation_rows
        if row["fiscal_year"] == last_year
        and row["scenario_id"] != row["baseline_scenario_id"]
    ]
    largest_interpretation = sorted(
        central_interpretation_last,
        key=lambda row: abs(
            _decimal(row["delta_central_ratewall_ratio_vs_baseline"])
        ),
        reverse=True,
    )[:5]
    largest_2036 = sorted(
        [
            row
            for row in normal
            if row["fiscal_year"] == last_year
            and row["scenario_id"] != "cbo_baseline_noop_v1"
        ],
        key=lambda row: abs(
            _decimal(row["delta_ratewall_ratio_vs_baseline_recomputed"])
        ),
        reverse=True,
    )[:5]
    class_counts = _classification_counts(channel_rows)
    numerator_plan_status_counts = _plan_status_counts(numerator_channel_plan_rows)
    zero_credit_relief_rows = [
        row for row in zero_low_apr_credit_rows if row["screen_relief_bil"]
    ]
    largest_zero_credit = sorted(
        zero_credit_relief_rows,
        key=lambda row: _decimal(row["screen_relief_bil"]),
        reverse=True,
    )[:3]
    deferred_channels_classified = len(channel_rows) - class_counts.get("included", 0)
    remaining_plan_rows = len(numerator_channel_plan_rows)
    settled_deferred_channels = max(0, deferred_channels_classified - remaining_plan_rows)
    public_interest_last = (
        _row_for(public_interest_rows, "cbo_baseline_noop_v1", last_year)
        if public_interest_rows
        else None
    )
    residual_last = [
        row
        for row in residual_sensitivity_rows
        if row["fiscal_year"] == last_year
        and row["scenario_id"] != row["baseline_scenario_id"]
    ]
    largest_residual = sorted(
        residual_last,
        key=lambda row: abs(
            _decimal(row["delta_total_residual_sensitivity_vs_baseline_bil"])
        ),
        reverse=True,
    )[:5]
    composition_last = [
        row
        for row in composition_surface_rows
        if row["fiscal_year"] == last_year
        and row["beta_path_id"] == "normal_forward_constant"
    ]
    largest_composition = sorted(
        [
            row
            for row in composition_last
            if row["scenario_id"] != row["baseline_scenario_id"]
        ],
        key=lambda row: abs(
            _decimal(row["delta_composition_ratewall_ratio_vs_baseline"])
        ),
        reverse=True,
    )[:5]
    baseline_current_composition = (
        _composition_row_for(
            composition_last,
            scenario_id="cbo_baseline_noop_v1",
            composition_case_id="first_forecast_current",
        )
        if composition_last
        else None
    )
    baseline_replacement_composition = (
        _composition_row_for(
            composition_last,
            scenario_id="cbo_baseline_noop_v1",
            composition_case_id="public_interest_replacement",
        )
        if composition_last
        else None
    )
    lines = [
        "# RateWall 10-Year Forecast Readout",
        "",
        "RateWall is `RW = N / D`. `N` is current-demand support. `D` is the conventional-demand shortfall. A scenario hits the wall when `RW >= 1`.",
        "",
        "## What Is Included",
        "",
        f"- The forecast uses the full-horizon TDCSim/CBO suite for FY{first_year}-FY{last_year}.",
        "- The first forecast numerator includes TDC ex-overlap support, direct Treasury interest support, and bank Treasury interest support.",
        "- TDC support is `TDC ex-overlap flow * beta * chi`.",
        f"- Baseline beta is `{DEFAULT_TDC_BETA}` and chi is `{DEFAULT_TDC_DEPOSIT_CURRENT_DEMAND_SHARE}`.",
        "- Rate-changing scenarios use the selected moving denominator already attached to the model-summary surface.",
        "",
        "## Baseline",
        "",
        f"- FY{first_year} baseline RW: `{baseline_first['level_ratewall_ratio_recomputed']}`.",
        f"- FY{last_year} baseline RW: `{baseline_last['level_ratewall_ratio_recomputed']}`.",
        f"- FY{last_year} baseline `N`: `{baseline_last['total_current_demand_support_bil_recomputed']}` billion.",
        f"- FY{last_year} baseline moving `D`: `{baseline_last['selected_moving_denominator_bil']}` billion.",
        "",
    ]
    if central_baseline_last is not None:
        strongest_positive = next(
            (
                row
                for row in largest_central
                if _decimal(row["delta_central_ratewall_ratio_vs_baseline"]) > 0
            ),
            None,
        )
        strongest_negative = next(
            (
                row
                for row in largest_central
                if _decimal(row["delta_central_ratewall_ratio_vs_baseline"]) < 0
            ),
            None,
        )
        lines.extend(
            [
                "## Preliminary Economist Summary",
                "",
                "- The central forecast is an Assumption-Mode scenario readout, not a canonical headline promotion.",
                "- Central `N` uses the public-interest replacement numerator and normal-forward TDC beta.",
                "- Rate scenarios move `D`; holder-route scenarios mainly move `N` through TDC routing.",
                f"- FY{last_year} central baseline RW is `{central_baseline_last['central_ratewall_ratio']}`.",
            ]
        )
        if strongest_positive is not None:
            lines.append(
                "- Largest positive central scenario move: "
                f"`{strongest_positive['scenario_id']}` changes RW by "
                f"`{strongest_positive['delta_central_ratewall_ratio_vs_baseline']}`."
            )
        if strongest_negative is not None:
            lines.append(
                "- Largest negative central scenario move: "
                f"`{strongest_negative['scenario_id']}` changes RW by "
                f"`{strongest_negative['delta_central_ratewall_ratio_vs_baseline']}`."
            )
        lines.extend(
            [
                "- Remaining numerator channels are treated as classified, sensitivity, replacement-only, or parked; none are silently assumed zero.",
                "",
                "## Central Forecast Choice",
                "",
                "- Central numerator: public-interest replacement.",
                "- Central beta path: normal-forward constant.",
                "- Residual and timed-beta alternatives remain sensitivity surfaces, not the central result.",
                f"- FY{last_year} central baseline RW: `{central_baseline_last['central_ratewall_ratio']}`.",
                f"- FY{last_year} central baseline `N`: `{central_baseline_last['central_n_bil']}` billion.",
                f"- FY{last_year} central moving `D`: `{central_baseline_last['central_moving_denominator_bil']}` billion.",
                "",
                f"## Largest FY{last_year} Central Scenario Moves",
                "",
            ]
        )
        for row in largest_central:
            lines.append(
                "- "
                f"`{row['scenario_id']}`: central delta RW "
                f"`{row['delta_central_ratewall_ratio_vs_baseline']}`, "
                f"delta `N` `{row['delta_central_n_vs_baseline_bil']}` bn, "
                f"delta `D` `{row['delta_central_moving_denominator_vs_baseline_bil']}` bn."
            )
        lines.append("")
    if largest_interpretation:
        lines.extend(
            [
                "## Plain Mechanism Readout",
                "",
            ]
        )
        for row in largest_interpretation:
            lines.append(
                "- "
                f"`{row['scenario_id']}` is `{row['scenario_direction']}` and "
                f"`{row['primary_driver']}`: {row['mechanism_summary']}. "
                f"Sensitivity width `{row['sensitivity_width_ratewall_ratio']}`."
        )
        lines.append("")
    if scenario_sufficiency_rows:
        active = sum(
            1
            for row in scenario_sufficiency_rows
            if row["in_central_surface"] == "true"
        )
        run_not_active = sum(
            1
            for row in scenario_sufficiency_rows
            if row["run_in_suite"] == "true" and row["in_central_surface"] == "false"
        )
        configured_not_run = sum(
            1
            for row in scenario_sufficiency_rows
            if row["configured_in_manifest"] == "true" and row["run_in_suite"] == "false"
        )
        required = [
            row
            for row in scenario_sufficiency_rows
            if row["coverage_status"] == "required_missing"
        ]
        holder_route_rows = [
            row
            for row in scenario_sufficiency_rows
            if row["scenario_axis"].startswith("holder_mix")
        ]
        lines.extend(
            [
                "## Scenario Coverage",
                "",
                f"- Active central scenario rows: `{active}` scenario IDs.",
                f"- Already run but not active: `{run_not_active}` scenario IDs.",
                f"- Configured but not run: `{configured_not_run}` scenario IDs.",
            ]
        )
        if holder_route_rows:
            lines.append(
                "- Holder-route coverage: current private-holder high/low rows are "
                "interpreted as reserve-user-like versus private/deposit-user shifts; "
                "`Banks`, `Foreign`, and `CB/Fed` are reserve-user-like for TDC route."
            )
            lines.append(
                "- Fed/CB reserve creation is a separate deferred channel; it is not "
                "added to the numerator in this readout."
            )
        for row in required:
            lines.append(
                f"- Required missing axis: `{row['scenario_axis']}`; "
                f"next action `{row['next_model_action']}`."
            )
        lines.append("")
    if public_interest_last is not None:
        lines.extend(
            [
                "## Public-Interest Net Block",
                "",
                "- The final interest block is staged as one replacement block, not an add-on to direct and bank interest.",
                f"- FY{last_year} baseline legacy direct+bank interest support: `{public_interest_last['legacy_interest_support_bil']}` billion.",
                f"- FY{last_year} baseline net support after Fed-liability/remittance additions and absorbers: `{public_interest_last['net_interest_after_fiscal_tga_offsets_bil']}` billion.",
                f"- FY{last_year} replacement delta vs legacy direct+bank support: `{public_interest_last['replacement_delta_vs_legacy_interest_support_bil']}` billion.",
                "- Foreign leakage is not re-applied because the current TDCSim direct/bank basis is already holder-filtered.",
                "",
            ]
        )
    if largest_residual:
        lines.extend(
            [
                "## Residual Numerator Sensitivities",
                "",
                "- Remaining firm-cash, household safe-yield, and deposit/MMF rows are staged as bounded sensitivities, not headline additions.",
                "- The basis is residual: it removes dollars already demand-converted in the public-interest block.",
                "- Deposit/MMF offset and drag are paired; the drag is flagged where it overlaps the moving denominator.",
            ]
        )
        for row in largest_residual:
            lines.append(
                "- "
                f"`{row['assumption_set']}` / `{row['scenario_id']}`: "
                f"FY{last_year} residual sensitivity delta "
                f"`{row['delta_total_residual_sensitivity_vs_baseline_bil']}` bn "
                f"(firm cash delta `{row['delta_firm_cash_attenuation_vs_baseline_bil']}`, "
                f"safe-yield delta `{row['delta_household_safe_yield_capture_vs_baseline_bil']}`, "
                f"paired deposit/MMF delta `{row['delta_paired_deposit_mmf_net_sensitivity_vs_baseline_bil']}`)."
            )
        lines.append("")
    if numerator_channel_plan_rows:
        lines.extend(
            [
                "## Remaining Numerator Channel Plan",
                "",
                f"- Remaining unresolved deferred channels: `{len(numerator_channel_plan_rows)}`.",
            ]
        )
        for status, count in sorted(numerator_plan_status_counts.items()):
            lines.append(f"- `{status}`: `{count}` channel(s).")
        lines.append(
            "- These rows are not added to central `N`; they define the route for admitting or parking each remaining channel without double counting."
        )
        for row in numerator_channel_plan_rows:
            lines.append(
                "- "
                f"`{row['channel_id']}`: `{row['final_central_status']}`; "
                f"next action `{row['next_model_action']}`."
            )
        lines.append("")
    if zero_low_apr_credit_rows:
        lines.extend(
            [
                "## Zero/Low-APR Credit Materiality Screen",
                "",
                f"- Product rows screened: `{len(zero_low_apr_credit_rows)}`.",
                "- This is a materiality screen only; no row enters central `N`.",
                "- The screen uses outstanding stock and duration where available; originations alone are blocked.",
            ]
        )
        for row in largest_zero_credit:
            lines.append(
                "- "
                f"`{row['product_segment']}`: screen relief "
                f"`{row['screen_relief_bil']}` bn; decision "
                f"`{row['materiality_decision']}`."
            )
        lines.append("")
    if (
        largest_composition
        and baseline_current_composition is not None
        and baseline_replacement_composition is not None
    ):
        lines.extend(
            [
                "## Forecast Composition Surface",
                "",
                "- The composition surface compares the first forecast numerator, the public-interest replacement numerator, and bounded residual alternatives against the selected moving denominator.",
                f"- FY{last_year} baseline first-forecast `N`: `{baseline_current_composition['composition_n_bil']}` bn; RW `{baseline_current_composition['composition_ratewall_ratio']}`.",
                f"- FY{last_year} baseline public-interest replacement `N`: `{baseline_replacement_composition['composition_n_bil']}` bn; RW `{baseline_replacement_composition['composition_ratewall_ratio']}`.",
            ]
        )
        for row in largest_composition:
            lines.append(
                "- "
                f"`{row['composition_case_id']}` / `{row['scenario_id']}`: "
                f"FY{last_year} delta RW "
                f"`{row['delta_composition_ratewall_ratio_vs_baseline']}`, "
                f"delta `N` `{row['delta_composition_n_vs_baseline_bil']}` bn."
            )
        lines.append("")
    lines.extend(
        [
            f"## Largest FY{last_year} Scenario Moves Under Normal-Forward Beta",
            "",
        ]
    )
    for row in largest_2036:
        lines.append(
            "- "
            f"`{row['scenario_id']}`: delta RW "
            f"`{row['delta_ratewall_ratio_vs_baseline_recomputed']}`, "
            f"delta `N` `{row['delta_total_current_demand_support_bil_recomputed']}` bn, "
            f"delta `D` `{row['selected_delta_denominator_bil']}` bn."
        )
    lines.extend(["", "## Timed Beta Paths", ""])
    beta_effects = _timed_beta_effect_summary(
        timed_beta_rows,
        fiscal_year=last_year,
    )
    for path_id, value in beta_effects:
        lines.append(
            f"- `{path_id}`: largest FY{last_year} RW change vs normal-forward path is `{value}`."
        )
    lines.extend(
        [
            "",
            "## Forecast Numerator Scope",
            "",
            f"- Included channels: `{class_counts.get('included', 0)}`.",
            f"- Deferred channels classified: `{deferred_channels_classified}`.",
            f"- Settled deferred channels outside the remaining-plan table: `{settled_deferred_channels}`.",
            f"- Remaining final-central admission/parking plan rows: `{remaining_plan_rows}`.",
            "- Deferred channels are classified as replacement-block pieces, bounded sensitivities, obsolete overlap rows, sidecars, or plan-gated rows; none are silently dropped.",
            "",
            "## Boundary",
            "",
            "- These are scenario-mode model artifacts, not canonical headline entries.",
            "- No beta prior, chi prior, denominator coefficient, MMF routing value, or canonical denominator formula is changed here.",
        ]
    )
    return "\n".join(lines) + "\n"


def _timed_beta_row(
    summary: Mapping[str, str],
    *,
    effect: Mapping[str, str],
    baseline_effect: Mapping[str, str],
    synthesis: Mapping[str, str],
    baseline_synthesis: Mapping[str, str],
    materiality: Mapping[str, str],
    beta_path: Mapping[str, str],
    beta_label: str,
    beta: Decimal,
    beta_source_status: str,
) -> dict[str, str]:
    chi = DEFAULT_TDC_DEPOSIT_CURRENT_DEMAND_SHARE
    scenario_support = _support(effect, beta=beta, chi=chi)
    baseline_support = _support(baseline_effect, beta=beta, chi=chi)
    denominator = _decimal(synthesis["selected_moving_denominator_bil"])
    baseline_denominator = _decimal(
        baseline_synthesis["selected_moving_denominator_bil"]
    )
    ratio = scenario_support["total"] / denominator
    baseline_ratio = baseline_support["total"] / baseline_denominator
    return {
        "forecast_timed_beta_path_row_id": (
            "forecast_timed_beta_path::"
            f"{beta_path['beta_path_id']}::{summary['fiscal_year']}::"
            f"{summary['scenario_id']}"
        ),
        "fiscal_year": summary["fiscal_year"],
        "scenario_id": summary["scenario_id"],
        "baseline_scenario_id": summary["baseline_scenario_id"],
        "beta_path_id": beta_path["beta_path_id"],
        "beta_path_label": beta_path["beta_path_label"],
        "beta_transition_fiscal_year": beta_path["transition_year"],
        "tdc_materialization_beta_scenario": beta_label,
        "tdc_materialization_beta": _fmt(beta),
        "tdc_materialization_beta_source_status": beta_source_status,
        "deposit_current_demand_share": _fmt(chi),
        "derived_beta_times_chi": _fmt(beta * chi),
        "tdc_change_ex_overlap_bil": effect["tdc_change_ex_overlap_bil"],
        "baseline_tdc_change_ex_overlap_bil": baseline_effect[
            "tdc_change_ex_overlap_bil"
        ],
        "tdc_current_demand_support_bil_recomputed": _fmt(scenario_support["tdc"]),
        "delta_tdc_current_demand_support_bil_recomputed": _fmt(
            scenario_support["tdc"] - baseline_support["tdc"]
        ),
        "direct_treasury_current_demand_support_bil_fixed": effect[
            "direct_treasury_current_demand_support_bil"
        ],
        "delta_direct_treasury_current_demand_support_bil_fixed": _fmt(
            scenario_support["direct"] - baseline_support["direct"]
        ),
        "bank_treasury_current_demand_support_bil_fixed": effect[
            "bank_treasury_current_demand_support_bil"
        ],
        "delta_bank_treasury_current_demand_support_bil_fixed": _fmt(
            scenario_support["bank"] - baseline_support["bank"]
        ),
        "total_current_demand_support_bil_recomputed": _fmt(
            scenario_support["total"]
        ),
        "delta_total_current_demand_support_bil_recomputed": _fmt(
            scenario_support["total"] - baseline_support["total"]
        ),
        "selected_moving_denominator_bil": synthesis[
            "selected_moving_denominator_bil"
        ],
        "selected_delta_denominator_bil": synthesis[
            "selected_delta_denominator_bil"
        ],
        "level_ratewall_ratio_recomputed": _fmt(ratio),
        "delta_ratewall_ratio_vs_baseline_recomputed": _fmt(
            ratio - baseline_ratio
        ),
        "delta_ratewall_ratio_vs_normal_forward_path": "0",
        "wall_hit_under_timed_beta_path": str(ratio >= Decimal("1")).lower(),
        "scenario_axis": materiality.get("scenario_family", summary["comparison_group"]),
        "model_relevance_class": materiality.get("model_relevance_class", ""),
        "allowed_use": "ten_year_forecast_timed_beta_scenario_readout",
        "blocked_use": (
            "canonical_headline_promotion;beta_prior_update;chi_prior_update;"
            "denominator_recalibration;evidence_mode_claim;release_headline_claim"
        ),
        "claim_boundary": (
            "scenario_mode_only;uses_existing_tdcsim_cbo_cashflows;"
            "beta_path_assumption_not_new_empirical_estimate"
        ),
        "canonical_ratio_entry": "false",
        "enters_main_ratio": "false",
        "evidence_mode_enabled": "false",
    }


def _composition_surface_row(
    timed_row: Mapping[str, str],
    *,
    public_interest: Mapping[str, str],
    composition_case_id: str,
    composition_case_label: str,
    residual_assumption_set: str,
    residual_sensitivity_delta: Decimal,
    use_public_interest_replacement: bool,
) -> dict[str, str]:
    tdc = _decimal(timed_row["tdc_current_demand_support_bil_recomputed"])
    direct = _decimal(timed_row["direct_treasury_current_demand_support_bil_fixed"])
    bank = _decimal(timed_row["bank_treasury_current_demand_support_bil_fixed"])
    legacy_interest = direct + bank
    public_interest_net = _decimal(
        public_interest["net_interest_after_fiscal_tga_offsets_bil"]
    )
    first_forecast_n = _decimal(timed_row["total_current_demand_support_bil_recomputed"])
    if use_public_interest_replacement:
        composition_n = tdc + public_interest_net + residual_sensitivity_delta
        composition_rule = (
            "public_interest_net_replaces_direct_plus_bank_interest;"
            "residual_sensitivity_enters_as_baseline_relative_delta"
        )
    else:
        composition_n = first_forecast_n
        composition_rule = "first_forecast_tdc_plus_direct_plus_bank_interest"
    denominator = _decimal(timed_row["selected_moving_denominator_bil"])
    ratio = composition_n / denominator
    return {
        "forecast_composition_surface_row_id": (
            "forecast_composition_surface::"
            f"{composition_case_id}::{timed_row['beta_path_id']}::"
            f"{timed_row['fiscal_year']}::{timed_row['scenario_id']}"
        ),
        "fiscal_year": timed_row["fiscal_year"],
        "scenario_id": timed_row["scenario_id"],
        "baseline_scenario_id": timed_row["baseline_scenario_id"],
        "beta_path_id": timed_row["beta_path_id"],
        "composition_case_id": composition_case_id,
        "composition_case_label": composition_case_label,
        "residual_assumption_set": residual_assumption_set,
        "tdc_current_demand_support_bil": _fmt(tdc),
        "direct_treasury_current_demand_support_bil": _fmt(direct),
        "bank_treasury_current_demand_support_bil": _fmt(bank),
        "legacy_interest_support_bil": _fmt(legacy_interest),
        "public_interest_net_support_bil": _fmt(public_interest_net),
        "interest_replacement_delta_bil": _fmt(public_interest_net - legacy_interest),
        "residual_sensitivity_delta_bil": _fmt(residual_sensitivity_delta),
        "first_forecast_n_bil": _fmt(first_forecast_n),
        "composition_n_bil": _fmt(composition_n),
        "baseline_composition_n_bil": "",
        "delta_composition_n_vs_baseline_bil": "",
        "selected_moving_denominator_bil": timed_row[
            "selected_moving_denominator_bil"
        ],
        "baseline_selected_moving_denominator_bil": "",
        "selected_delta_denominator_bil": timed_row["selected_delta_denominator_bil"],
        "composition_ratewall_ratio": _fmt(ratio),
        "baseline_composition_ratewall_ratio": "",
        "delta_composition_ratewall_ratio_vs_baseline": "",
        "wall_hit_under_composition": str(ratio >= Decimal("1")).lower(),
        "composition_rule": composition_rule,
        "allowed_use": "forecast_model_composition_scenario_surface",
        "blocked_use": (
            "canonical_headline_promotion;release_headline_claim;"
            "evidence_mode_claim;denominator_prior_update;beta_prior_update;"
            "chi_prior_update"
        ),
        "canonical_ratio_entry": "false",
        "enters_main_ratio": "false",
        "evidence_mode_enabled": "false",
    }


def _add_composition_delta_fields(
    row: dict[str, str],
    baseline: Mapping[str, str],
) -> None:
    baseline_n = _decimal(baseline["composition_n_bil"])
    baseline_d = _decimal(baseline["selected_moving_denominator_bil"])
    baseline_ratio = _decimal(baseline["composition_ratewall_ratio"])
    row["baseline_composition_n_bil"] = _fmt(baseline_n)
    row["delta_composition_n_vs_baseline_bil"] = _fmt(
        _decimal(row["composition_n_bil"]) - baseline_n
    )
    row["baseline_selected_moving_denominator_bil"] = _fmt(baseline_d)
    row["baseline_composition_ratewall_ratio"] = _fmt(baseline_ratio)
    row["delta_composition_ratewall_ratio_vs_baseline"] = _fmt(
        _decimal(row["composition_ratewall_ratio"]) - baseline_ratio
    )


def _primary_driver(
    numerator_only: Decimal,
    denominator_only: Decimal,
) -> str:
    n_abs = abs(numerator_only)
    d_abs = abs(denominator_only)
    if n_abs == 0 and d_abs == 0:
        return "baseline_or_zero_delta"
    if d_abs == 0:
        return "numerator_driven"
    if n_abs == 0:
        return "denominator_driven"
    ratio = n_abs / d_abs
    if Decimal("0.67") <= ratio <= Decimal("1.5"):
        return "mixed_numerator_and_denominator"
    return "numerator_driven" if n_abs > d_abs else "denominator_driven"


def _scenario_direction(delta_ratio: Decimal) -> str:
    if delta_ratio > 0:
        return "raises_ratewall"
    if delta_ratio < 0:
        return "lowers_ratewall"
    return "no_ratewall_change"


def _mechanism_summary(
    *,
    delta_n: Decimal,
    delta_d: Decimal,
    numerator_only: Decimal,
    denominator_only: Decimal,
    primary_driver: str,
) -> str:
    n_text = (
        "higher N raises RW"
        if delta_n > 0
        else "lower N lowers RW"
        if delta_n < 0
        else "N unchanged"
    )
    d_text = (
        "higher D lowers RW"
        if delta_d > 0
        else "lower D raises RW"
        if delta_d < 0
        else "D unchanged"
    )
    return (
        f"{primary_driver}; {n_text}; {d_text}; "
        f"N-only delta {_fmt(numerator_only)}, D-only delta {_fmt(denominator_only)}"
    )


def _support(
    effect: Mapping[str, str],
    *,
    beta: Decimal,
    chi: Decimal,
) -> dict[str, Decimal]:
    tdc = _decimal(effect["tdc_change_ex_overlap_bil"]) * beta * chi
    direct = _decimal(effect["direct_treasury_current_demand_support_bil"])
    bank = _decimal(effect["bank_treasury_current_demand_support_bil"])
    return {"tdc": tdc, "direct": direct, "bank": bank, "total": tdc + direct + bank}


def _beta_for_path(
    beta_path: Mapping[str, str],
    fiscal_year: int,
) -> tuple[str, Decimal, str]:
    transition = beta_path["transition_year"]
    beta_label = beta_path["early_beta"]
    if transition and fiscal_year >= int(transition):
        beta_label = beta_path["late_beta"]
    try:
        beta, source_status = _BETA_PROFILE_BY_ID[beta_label]
    except KeyError as exc:
        raise ForecastModelReadoutError(f"unknown beta profile: {beta_label}") from exc
    return beta_label, beta, source_status


def _write_pngs(
    paths: Mapping[str, Path],
    timed_beta_rows: Sequence[Mapping[str, str]],
    channel_rows: Sequence[Mapping[str, str]],
    composition_surface_rows: Sequence[Mapping[str, str]],
    central_forecast_rows: Sequence[Mapping[str, str]],
    central_interpretation_rows: Sequence[Mapping[str, str]],
    scenario_sufficiency_rows: Sequence[Mapping[str, str]],
) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    normal = [
        row
        for row in timed_beta_rows
        if row["beta_path_id"] == "normal_forward_constant"
    ]
    top_2036 = sorted(
        [
            row
            for row in normal
            if row["fiscal_year"] == "2036"
            and row["scenario_id"] != "cbo_baseline_noop_v1"
        ],
        key=lambda row: abs(
            _decimal(row["delta_ratewall_ratio_vs_baseline_recomputed"])
        ),
        reverse=True,
    )[:4]
    selected_ids = ["cbo_baseline_noop_v1", *(row["scenario_id"] for row in top_2036)]
    fig, ax = plt.subplots(figsize=(11, 6.2))
    for scenario_id in selected_ids:
        series = sorted(
            [
                row
                for row in normal
                if row["scenario_id"] == scenario_id
            ],
            key=lambda row: int(row["fiscal_year"]),
        )
        ax.plot(
            [int(row["fiscal_year"]) for row in series],
            [_float(row["level_ratewall_ratio_recomputed"]) for row in series],
            marker="o",
            label=_short_label(scenario_id),
        )
    ax.axhline(1, color="#111827", linewidth=0.8, linestyle="--")
    ax.set_title("10-Year RateWall Paths under Normal-Forward Beta")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("RateWall ratio")
    ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(paths["png_ratewall_paths"], dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(10.5, 5.8))
    for path_id in (
        "latest_rolling_from_fy2032",
        "pooled_full_from_fy2032_review",
    ):
        rows = [
            row
            for row in timed_beta_rows
            if row["beta_path_id"] == path_id
            and row["scenario_id"] != "cbo_baseline_noop_v1"
        ]
        by_year: dict[int, float] = {}
        for year in sorted({int(row["fiscal_year"]) for row in rows}):
            year_rows = [row for row in rows if int(row["fiscal_year"]) == year]
            by_year[year] = max(
                abs(_float(row["delta_ratewall_ratio_vs_normal_forward_path"]))
                for row in year_rows
            )
        ax.plot(by_year.keys(), by_year.values(), marker="o", label=path_id)
    ax.set_title("Timed Beta Paths: Largest Absolute RW Change vs Normal-Forward")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("Absolute RateWall ratio change")
    ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(paths["png_timed_beta_effect"], dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(11, 6.2))
    width = 0.24
    x = range(len(top_2036))
    components = [
        ("delta_tdc_current_demand_support_bil_recomputed", "TDC", "#2563eb"),
        (
            "delta_direct_treasury_current_demand_support_bil_fixed",
            "Direct interest",
            "#7c3aed",
        ),
        (
            "delta_bank_treasury_current_demand_support_bil_fixed",
            "Bank interest",
            "#0891b2",
        ),
        ("selected_delta_denominator_bil", "D move", "#ea580c"),
    ]
    offsets = (-1.5 * width, -0.5 * width, 0.5 * width, 1.5 * width)
    for offset, (field, label, color) in zip(offsets, components, strict=True):
        ax.bar(
            [index + offset for index in x],
            [_float(row[field]) for row in top_2036],
            width,
            label=label,
            color=color,
        )
    ax.axhline(0, color="#111827", linewidth=0.8)
    ax.set_xticks(list(x), [_short_label(row["scenario_id"]) for row in top_2036])
    ax.tick_params(axis="x", rotation=25)
    ax.set_title("FY2036 Largest Scenarios: N Components and D Move")
    ax.set_ylabel("Billion dollars")
    ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(paths["png_components"], dpi=180)
    plt.close(fig)

    counts = _classification_counts(channel_rows)
    fig, ax = plt.subplots(figsize=(7.5, 4.8))
    labels = ["included", "still missing"]
    values = [
        counts.get("included", 0),
        counts.get("still_missing_deferred", 0),
    ]
    ax.bar(labels, values, color=["#2563eb", "#dc2626"])
    ax.set_title("Forecast Numerator Channel Scope")
    ax.set_ylabel("Channel count")
    for index, value in enumerate(values):
        ax.text(index, value + 0.2, str(value), ha="center")
    fig.tight_layout()
    fig.savefig(paths["png_channel_scope"], dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(11, 6.2))
    composition_last = [
        row
        for row in composition_surface_rows
        if row["fiscal_year"] == "2036"
        and row["beta_path_id"] == "normal_forward_constant"
        and row["scenario_id"] != row["baseline_scenario_id"]
    ]
    case_ids = [
        "first_forecast_current",
        "public_interest_replacement",
        "public_interest_plus_residual_delta::literature_calibrated_base",
        (
            "public_interest_plus_residual_delta::"
            "assumption_mode_deposit_mmf_paired_entry"
        ),
    ]
    if composition_last:
        top_ids = [
            row["scenario_id"]
            for row in sorted(
                [
                    row
                    for row in composition_last
                    if row["composition_case_id"] == "public_interest_replacement"
                ],
                key=lambda row: abs(
                    _decimal(row["delta_composition_ratewall_ratio_vs_baseline"])
                ),
                reverse=True,
            )[:4]
        ]
        width = 0.18
        x = range(len(top_ids))
        offsets = (-1.5 * width, -0.5 * width, 0.5 * width, 1.5 * width)
        colors = ("#2563eb", "#0891b2", "#7c3aed", "#ea580c")
        for offset, case_id, color in zip(offsets, case_ids, colors, strict=True):
            values = [
                _float(
                    _composition_row_for(
                        composition_last,
                        scenario_id=scenario_id,
                        composition_case_id=case_id,
                    )["delta_composition_ratewall_ratio_vs_baseline"]
                )
                for scenario_id in top_ids
            ]
            ax.bar(
                [index + offset for index in x],
                values,
                width,
                label=_short_composition_case(case_id),
                color=color,
            )
        ax.set_xticks(list(x), [_short_label(scenario_id) for scenario_id in top_ids])
        ax.tick_params(axis="x", rotation=25)
    else:
        ax.text(0.5, 0.5, "No composition rows", ha="center", va="center")
        ax.set_xticks([])
    ax.axhline(0, color="#111827", linewidth=0.8)
    ax.set_title("FY2036 Composition Surface: Delta RateWall Ratio")
    ax.set_ylabel("Delta RateWall ratio vs baseline")
    if composition_last:
        ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(paths["png_composition_surface"], dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(11, 6.2))
    central_last = [
        row
        for row in central_forecast_rows
        if row["fiscal_year"] == "2036"
        and row["scenario_id"] != row["baseline_scenario_id"]
    ]
    if central_last:
        top_rows = sorted(
            central_last,
            key=lambda row: abs(
                _decimal(row["delta_central_ratewall_ratio_vs_baseline"])
            ),
            reverse=True,
        )[:6]
        labels = [_short_label(row["scenario_id"]) for row in top_rows]
        values = [
            _float(row["delta_central_ratewall_ratio_vs_baseline"])
            for row in top_rows
        ]
        colors = ["#2563eb" if value >= 0 else "#ea580c" for value in values]
        x = range(len(top_rows))
        ax.bar(x, values, color=colors)
        ax.set_xticks(list(x), labels)
        ax.tick_params(axis="x", rotation=25)
    else:
        ax.text(0.5, 0.5, "No central forecast rows", ha="center", va="center")
        ax.set_xticks([])
    ax.axhline(0, color="#111827", linewidth=0.8)
    ax.set_title("FY2036 Central Forecast: Delta RateWall Ratio")
    ax.set_ylabel("Delta RateWall ratio vs baseline")
    fig.tight_layout()
    fig.savefig(paths["png_central_surface"], dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(11, 6.2))
    baseline_series = sorted(
        [
            row
            for row in central_forecast_rows
            if row["scenario_id"] == row["baseline_scenario_id"]
        ],
        key=lambda row: int(row["fiscal_year"]),
    )
    if baseline_series:
        ax.plot(
            [int(row["fiscal_year"]) for row in baseline_series],
            [_float(row["central_ratewall_ratio"]) for row in baseline_series],
            marker="o",
            color="#2563eb",
            label="central baseline",
        )
        ax.axhline(1, color="#111827", linewidth=0.8, linestyle="--")
        ax.legend(fontsize=8)
    else:
        ax.text(0.5, 0.5, "No central baseline rows", ha="center", va="center")
        ax.set_xticks([])
    ax.set_title("Central Forecast Baseline Path")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("RateWall ratio")
    fig.tight_layout()
    fig.savefig(paths["png_central_baseline_path"], dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(11, 6.2))
    interpretation_last = [
        row
        for row in central_interpretation_rows
        if row["fiscal_year"] == "2036"
        and row["scenario_id"] != row["baseline_scenario_id"]
    ]
    if interpretation_last:
        top_rows = sorted(
            interpretation_last,
            key=lambda row: abs(
                _decimal(row["delta_central_ratewall_ratio_vs_baseline"])
            ),
            reverse=True,
        )[:6]
        labels = [_short_label(row["scenario_id"]) for row in top_rows]
        lows = [_float(row["sensitivity_min_ratewall_ratio"]) for row in top_rows]
        highs = [_float(row["sensitivity_max_ratewall_ratio"]) for row in top_rows]
        centers = [_float(row["central_ratewall_ratio"]) for row in top_rows]
        x = list(range(len(top_rows)))
        lower_err = [center - low for center, low in zip(centers, lows, strict=True)]
        upper_err = [high - center for high, center in zip(highs, centers, strict=True)]
        ax.errorbar(
            x,
            centers,
            yerr=[lower_err, upper_err],
            fmt="o",
            color="#2563eb",
            ecolor="#64748b",
            capsize=5,
            label="central with sensitivity range",
        )
        ax.set_xticks(x, labels)
        ax.tick_params(axis="x", rotation=25)
        ax.legend(fontsize=8)
    else:
        ax.text(0.5, 0.5, "No interpretation rows", ha="center", va="center")
        ax.set_xticks([])
    ax.set_title("FY2036 Central Forecast Sensitivity Range")
    ax.set_ylabel("RateWall ratio")
    fig.tight_layout()
    fig.savefig(paths["png_central_sensitivity_spread"], dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(9.5, 5.6))
    sufficiency_counts = _scenario_sufficiency_counts(scenario_sufficiency_rows)
    labels = [
        "active",
        "run not active",
        "configured not run",
        "required missing",
    ]
    keys = [
        "active_central_surface",
        "run_not_active",
        "configured_not_run",
        "required_missing",
    ]
    values = [sufficiency_counts.get(key, 0) for key in keys]
    colors = ["#2563eb", "#0891b2", "#7c3aed", "#dc2626"]
    x = range(len(labels))
    ax.bar(x, values, color=colors)
    ax.set_xticks(list(x), labels)
    ax.set_title("Scenario Coverage for Current Forecast Readout")
    ax.set_ylabel("Scenario or required-axis count")
    for index, value in enumerate(values):
        ax.text(index, value + 0.15, str(value), ha="center")
    fig.tight_layout()
    fig.savefig(paths["png_scenario_sufficiency"], dpi=180)
    plt.close(fig)


def _assumption_set(name: str):
    for assumption in load_ratewall_assumption_sets():
        if assumption.name == name:
            return assumption
    raise ForecastModelReadoutError(f"unknown assumption set: {name}")


def _read_fred_csv(path: Path) -> list[tuple[date, Decimal]]:
    observations: list[tuple[date, Decimal]] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = reader.fieldnames or []
        if len(fields) < 2 or fields[0] != "observation_date":
            raise ForecastModelReadoutError(f"unexpected FRED CSV header: {path}")
        value_field = fields[1]
        for row in reader:
            value = (row.get(value_field) or "").strip()
            if not value or value == ".":
                continue
            observations.append(
                (date.fromisoformat(row["observation_date"]), _decimal(value))
            )
    return observations


def _cbo_sheet_values(sheet, row_label: str) -> list[object]:
    for row in sheet.iter_rows(values_only=True):
        if str(row[0]).strip() == row_label:
            return list(row[2:])
    raise ForecastModelReadoutError(f"missing CBO row: {row_label}")


def _cbo_sheet_year_values(sheet) -> list[object]:
    for row in sheet.iter_rows(values_only=True):
        if row[1] == "Units":
            return list(row[2:])
    raise ForecastModelReadoutError("missing CBO fiscal-year header row")


def _year_from_value(value: object) -> int | None:
    if value is None:
        return None
    try:
        year = int(str(value).strip())
    except ValueError:
        return None
    return year if 1900 <= year <= 2200 else None


def _stock_gdp_share(
    source_row: Mapping[str, str],
    macro_by_year: Mapping[str, Mapping[str, str]],
    label: str,
) -> Decimal:
    first_year = str(min(int(year) for year in macro_by_year))
    gdp = _decimal(macro_by_year[first_year]["cbo_nominal_gdp_bil"])
    if gdp <= 0:
        raise ForecastModelReadoutError(f"nonpositive CBO GDP for {first_year}")
    stock_bil = _millions_to_bil(_decimal(source_row["latest_average_value"]))
    if stock_bil < 0:
        raise ForecastModelReadoutError(f"{label} stock must not be negative")
    return stock_bil / gdp


def _rate_spread(
    source_row: Mapping[str, str] | None,
    macro_by_year: Mapping[str, Mapping[str, str]],
    *,
    missing_status: str,
) -> tuple[Decimal, str]:
    if source_row is None:
        return Decimal("0"), missing_status
    first_year = str(min(int(year) for year in macro_by_year))
    short_rate = _decimal(macro_by_year[first_year]["cbo_short_rate_pct"])
    spread = _decimal(source_row["latest_average_value"]) - short_rate
    return spread, "source_backed_latest_rate_spread_held_constant"


def _firm_liquid_asset_stock_share(
    source_rows: Sequence[Mapping[str, str]],
    macro_by_year: Mapping[str, Mapping[str, str]],
    fallback_assumption,
) -> tuple[Decimal, str]:
    source_by_series = {row["series_id"]: row for row in source_rows}
    missing = [
        series_id
        for series_id in RESIDUAL_SENSITIVITY_FRED_SERIES
        if series_id not in source_by_series
    ]
    if missing:
        return (
            _decimal(fallback_assumption.firm_liquid_asset_stock_share_gdp),
            "fallback_assumption_missing_fred_components:"
            + ";".join(sorted(missing)),
        )
    first_year = str(min(int(year) for year in macro_by_year))
    gdp = _decimal(macro_by_year[first_year]["cbo_nominal_gdp_bil"])
    if gdp <= 0:
        raise ForecastModelReadoutError(f"nonpositive CBO GDP for {first_year}")
    stock_bil = sum(
        (
            _millions_to_bil(_decimal(row["latest_average_value"]))
            for row in source_by_series.values()
        ),
        Decimal("0"),
    )
    if stock_bil < 0:
        raise ForecastModelReadoutError("firm liquid asset stock must not be negative")
    return (
        stock_bil / gdp,
        "source_backed_latest_4q_average_z1_fred_components_held_constant",
    )


def _add_residual_delta_fields(
    row: dict[str, str],
    baseline: Mapping[str, str],
) -> None:
    component_pairs = (
        (
            "firm_cash_attenuation_bil",
            "baseline_firm_cash_attenuation_bil",
            "delta_firm_cash_attenuation_vs_baseline_bil",
        ),
        (
            "household_safe_yield_capture_bil",
            "baseline_household_safe_yield_capture_bil",
            "delta_household_safe_yield_capture_vs_baseline_bil",
        ),
        (
            "paired_deposit_mmf_net_sensitivity_bil",
            "baseline_paired_deposit_mmf_net_sensitivity_bil",
            "delta_paired_deposit_mmf_net_sensitivity_vs_baseline_bil",
        ),
        (
            "total_residual_sensitivity_bil",
            "baseline_total_residual_sensitivity_bil",
            "delta_total_residual_sensitivity_vs_baseline_bil",
        ),
    )
    for value_field, baseline_field, delta_field in component_pairs:
        base_value = _decimal(baseline[value_field])
        row[baseline_field] = _fmt(base_value)
        row[delta_field] = _fmt(_decimal(row[value_field]) - base_value)


def _millions_to_bil(value: Decimal) -> Decimal:
    return value / Decimal("1000")


def _suite_files(suite_dir: str | Path) -> _SuiteFiles:
    root = Path(suite_dir)
    artifact = ArtifactManifestView.from_root(root) if artifact_manifest_exists(root) else None
    return _SuiteFiles(root=root, artifact=artifact)


def _read_csv(files: _SuiteFiles, logical_path: str) -> list[dict[str, str]]:
    if files.artifact is not None:
        if not files.artifact.has_file(logical_path):
            raise ForecastModelReadoutError(f"missing suite CSV: {logical_path}")
        with files.artifact.open_text(logical_path) as handle:
            return list(csv.DictReader(handle))
    path = files.root / logical_path
    if not path.exists():
        raise ForecastModelReadoutError(f"missing suite CSV: {path}")
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_csv(
    path: Path,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, str]],
) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _by_key(
    rows: Iterable[Mapping[str, str]],
    label: str,
    *,
    require_unique: bool = True,
) -> dict[tuple[str, str], dict[str, str]]:
    out: dict[tuple[str, str], dict[str, str]] = {}
    for row in rows:
        key = (row["scenario_id"], row["fiscal_year"])
        if require_unique and key in out:
            raise ForecastModelReadoutError(f"duplicate {label} row for {key}")
        out[key] = dict(row)
    return out


def _by_key_with_optional_axis(
    rows: Iterable[Mapping[str, str]],
    label: str,
    *,
    axis_field: str,
) -> dict[tuple[str, str, str], dict[str, str]]:
    out: dict[tuple[str, str, str], dict[str, str]] = {}
    for row in rows:
        key = (row["scenario_id"], row["fiscal_year"], row[axis_field])
        if key in out:
            raise ForecastModelReadoutError(f"duplicate {label} row for {key}")
        out[key] = dict(row)
    return out


def _required(
    rows: Mapping[tuple[str, str], dict[str, str]],
    key: tuple[str, str],
    label: str,
) -> dict[str, str]:
    try:
        return rows[key]
    except KeyError as exc:
        raise ForecastModelReadoutError(f"missing {label} row for {key}") from exc


def _required_composition_case(
    rows: Mapping[tuple[str, str, str, str], dict[str, str]],
    *,
    case_id: str,
    beta_path_id: str,
    scenario_id: str,
    fiscal_year: str,
) -> dict[str, str]:
    key = (case_id, beta_path_id, scenario_id, fiscal_year)
    try:
        return rows[key]
    except KeyError as exc:
        raise ForecastModelReadoutError(
            f"missing composition sensitivity row for {key}"
        ) from exc


def _row_for(
    rows: Sequence[Mapping[str, str]],
    scenario_id: str,
    fiscal_year: str,
) -> Mapping[str, str]:
    for row in rows:
        if row["scenario_id"] == scenario_id and row["fiscal_year"] == fiscal_year:
            return row
    raise ForecastModelReadoutError(f"missing row for {scenario_id}::{fiscal_year}")


def _composition_row_for(
    rows: Sequence[Mapping[str, str]],
    *,
    scenario_id: str,
    composition_case_id: str,
) -> Mapping[str, str]:
    for row in rows:
        if (
            row["scenario_id"] == scenario_id
            and row["composition_case_id"] == composition_case_id
        ):
            return row
    raise ForecastModelReadoutError(
        f"missing composition row for {scenario_id}::{composition_case_id}"
    )


def _timed_beta_effect_summary(
    rows: Sequence[Mapping[str, str]],
    *,
    fiscal_year: str,
) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for path_id in (
        "latest_rolling_from_fy2032",
        "pooled_full_from_fy2032_review",
    ):
        values = [
            abs(_decimal(row["delta_ratewall_ratio_vs_normal_forward_path"]))
            for row in rows
            if row["beta_path_id"] == path_id
            and row["fiscal_year"] == fiscal_year
            and row["scenario_id"] != "cbo_baseline_noop_v1"
        ]
        out.append((path_id, _fmt(max(values) if values else Decimal("0"))))
    return out


def _classification_counts(rows: Sequence[Mapping[str, str]]) -> dict[str, int]:
    out: dict[str, int] = {}
    for row in rows:
        out[row["classification"]] = out.get(row["classification"], 0) + 1
    return out


def _plan_status_counts(rows: Sequence[Mapping[str, str]]) -> dict[str, int]:
    out: dict[str, int] = {}
    for row in rows:
        out[row["final_central_status"]] = out.get(row["final_central_status"], 0) + 1
    return out


def _scenario_config_review_rows(files: _SuiteFiles) -> list[dict[str, str]]:
    if files.artifact is not None:
        logical_paths = files.artifact.list_files(prefix="scenarios/", suffix=".json")
        payloads = [
            json.loads(files.artifact.read_text(logical_path))
            for logical_path in logical_paths
        ]
    else:
        scenario_dir = files.root / "scenarios"
        payloads = (
            [
                json.loads(path.read_text(encoding="utf-8"))
                for path in sorted(scenario_dir.glob("*.json"))
            ]
            if scenario_dir.exists()
            else []
        )
    rows: list[dict[str, str]] = []
    for payload in payloads:
        if not isinstance(payload, Mapping):
            continue
        scenario_id = str(payload.get("scenario_id", ""))
        if not scenario_id:
            continue
        provenance = payload.get("provenance", {})
        if not isinstance(provenance, Mapping):
            provenance = {}
        rows.append(
            {
                "scenario_id": scenario_id,
                "scenario_title": str(payload.get("title", "")),
                "scenario_axis": _scenario_axis_from_config(payload),
                "provenance_kind": str(provenance.get("kind", "")),
            }
        )
    return sorted(rows, key=lambda row: row["scenario_id"])


def _scenario_axis_from_config(payload: Mapping[str, object]) -> str:
    scenario_id = str(payload.get("scenario_id", ""))
    overrides = payload.get("overrides", {})
    if not isinstance(overrides, Mapping):
        overrides = {}
    if scenario_id == "cbo_baseline_noop_v1":
        return "baseline"
    if scenario_id in {
        "tdcsim_private_holder_high_v1",
        "tdcsim_private_holder_low_v1",
    }:
        return "holder_mix_reserve_user_private"
    if "holder_source" in scenario_id:
        return "holder_mix_reserve_user"
    if "mmf_pass_through" in scenario_id:
        return "mmf_routing"
    if "primary_deficit" in scenario_id:
        return "primary_deficit"
    if "rate_" in scenario_id or (
        "nominal_yield_curve" in overrides and "issuance_mix" not in overrides
    ):
        return "rate_curve"
    if "empirical" in scenario_id and "termprem" in scenario_id:
        return "issuance_term_premium"
    if "empirical" in scenario_id and "issuance" in scenario_id:
        return "issuance_empirical_mix"
    if scenario_id in {"tdcsim_issuance_shorter_v1", "tdcsim_issuance_longer_v1"}:
        return "issuance_generic_superseded"
    if "issuance_mix" in overrides:
        return "issuance_mix"
    if any(
        key in overrides for key in ("fed_holdings", "operating_cash", "fiscal_incidence")
    ):
        return "fiscal_fed_cash"
    return "other"


def _holder_route_taxonomy_note() -> str:
    ru = "+".join(RESERVE_USER_TDC_ROUTE_HOLDER_TYPES)
    private = "+".join(PRIVATE_DEPOSIT_USER_TDC_ROUTE_HOLDER_TYPES)
    return (
        f"tdc_route_taxonomy:{ru}=reserve_user_like;{private}=private_deposit_user_like;"
        "fed_cb_reserve_creation_channel_deferred"
    )


def _scenario_sufficiency_row(
    *,
    scenario_id: str,
    config: Mapping[str, str],
    run_in_suite: bool,
    central_2036: Mapping[str, str] | None,
) -> dict[str, str]:
    configured = bool(config) and not scenario_id.startswith("required_")
    in_central = central_2036 is not None
    axis = config.get("scenario_axis") or _scenario_axis_from_id(scenario_id)
    coverage_status = _scenario_coverage_status(
        configured=configured,
        run_in_suite=run_in_suite,
        in_central=in_central,
        scenario_id=scenario_id,
    )
    decision, action = _scenario_sufficiency_decision(
        scenario_id=scenario_id,
        axis=axis,
        coverage_status=coverage_status,
    )
    return {
        "forecast_scenario_sufficiency_row_id": (
            f"forecast_scenario_sufficiency::{scenario_id}"
        ),
        "scenario_id": scenario_id,
        "scenario_title": config.get("scenario_title", ""),
        "scenario_axis": axis,
        "configured_in_manifest": str(configured).lower(),
        "run_in_suite": str(run_in_suite).lower(),
        "in_central_surface": str(in_central).lower(),
        "provenance_kind": config.get("provenance_kind", ""),
        "tdc_route_taxonomy": _holder_route_taxonomy_note()
        if axis.startswith("holder_mix")
        else "",
        "coverage_status": coverage_status,
        "sufficiency_decision": decision,
        "next_model_action": action,
        "fy2036_delta_central_ratewall_ratio": (
            central_2036["delta_central_ratewall_ratio_vs_baseline"]
            if central_2036 is not None
            else ""
        ),
        "fy2036_primary_driver": (
            central_2036["primary_driver"] if central_2036 is not None else ""
        ),
        "allowed_use": "forecast_scenario_sufficiency_review",
        "blocked_use": (
            "canonical_headline_promotion;release_headline_claim;"
            "scenario_claim_without_run_row;scenario_claim_without_model_surface"
        ),
        "canonical_ratio_entry": "false",
        "enters_main_ratio": "false",
        "evidence_mode_enabled": "false",
    }


def _scenario_axis_from_id(scenario_id: str) -> str:
    return _scenario_axis_from_config({"scenario_id": scenario_id, "overrides": {}})


def _scenario_coverage_status(
    *,
    configured: bool,
    run_in_suite: bool,
    in_central: bool,
    scenario_id: str,
) -> str:
    if scenario_id.startswith("required_"):
        return "required_missing"
    if in_central:
        return "active_central_surface"
    if run_in_suite:
        return "run_not_active"
    if configured:
        return "configured_not_run"
    return "unknown_not_configured"


def _scenario_sufficiency_decision(
    *,
    scenario_id: str,
    axis: str,
    coverage_status: str,
) -> tuple[str, str]:
    if coverage_status == "active_central_surface":
        if axis == "holder_mix_reserve_user_private":
            return (
                "covered_as_reserve_user_vs_private_tdc_route_shift",
                "interpret_as_reserve_user_like_absorption_shift_with_fed_reserve_creation_deferred",
            )
        return "covered_in_current_central_surface", "interpret_in_current_readout"
    if scenario_id in {"tdcsim_issuance_shorter_v1", "tdcsim_issuance_longer_v1"}:
        return (
            "superseded_by_empirical_issuance_surface",
            "do_not_promote_unless_empirical_surface_is_withdrawn",
        )
    if axis == "issuance_term_premium" and coverage_status == "configured_not_run":
        return (
            "useful_bound_not_needed_for_central_point",
            "run_conservative_high_bounds_only_if_error_bands_become_next_model_focus",
        )
    if axis == "primary_deficit" and coverage_status == "configured_not_run":
        return (
            "scale_comparator_available_but_not_current_scenario_focus",
            "run_if_economist_readout_needs_deficit_scale_benchmark",
        )
    if axis == "fiscal_fed_cash" and coverage_status == "configured_not_run":
        return (
            "defer_until_public_interest_block_or_cash_policy_changes",
            "do_not_run_before_holder_and_rate_coverage",
        )
    return (
        "not_selected_for_current_central_surface",
        "review_only_do_not_promote_without_model_decision",
    )


def _scenario_sufficiency_sort(row: Mapping[str, str]) -> int:
    order = {
        "active_central_surface": 0,
        "run_not_active": 1,
        "configured_not_run": 2,
        "required_missing": 3,
    }
    return order.get(row["coverage_status"], 9)


def _scenario_sufficiency_counts(
    rows: Sequence[Mapping[str, str]],
) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["coverage_status"]] = counts.get(row["coverage_status"], 0) + 1
    return counts


def _short_label(scenario_id: str) -> str:
    label = scenario_id.replace("tdcsim_", "").replace("_v1", "")
    label = label.replace("issuance_empirical_", "")
    label = label.replace("termprem_", "tp_")
    return label[:42]


def _short_composition_case(case_id: str) -> str:
    if case_id == "first_forecast_current":
        return "first N"
    if case_id == "public_interest_replacement":
        return "replacement N"
    if case_id.endswith("literature_calibrated_base"):
        return "replacement + residual base"
    if case_id.endswith("assumption_mode_deposit_mmf_paired_entry"):
        return "replacement + residual paired"
    return case_id[:32]


def _decimal(value: object) -> Decimal:
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ForecastModelReadoutError(f"invalid decimal value: {value}") from exc


def _optional_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    return _decimal(value)


def _float(value: object) -> float:
    return float(_decimal(value))


def _fmt(value: Decimal) -> str:
    if value == value.to_integral():
        return str(value.quantize(Decimal("1")))
    return format(value.normalize(), "f")
