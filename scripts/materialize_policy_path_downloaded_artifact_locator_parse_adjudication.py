"""Parse downloaded policy-path artifacts into review-only locator candidates."""

from __future__ import annotations

import csv
import hashlib
import re
import sys
import zipfile
from html import unescape
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
INPUT = (
    ROOT
    / "outputs/tables/ratewall_policy_path_real_source_author_web_acquisition_attempt_packet.csv"
)
OUT_DIR = ROOT / "data/raw/policy_path_downloaded_artifact_locator_parse_adjudication"
MANIFEST = OUT_DIR / "policy_path_downloaded_artifact_locator_parse_adjudication_manifest.csv"

FIELDS = [
    "policy_path_downloaded_artifact_locator_parse_adjudication_manifest_row_id",
    "policy_path_real_source_author_web_acquisition_attempt_packet_row_id",
    "authored_field_name",
    "parse_attempt_class",
    "candidate_locator_count",
    "candidate_source_artifact_paths",
    "candidate_source_artifact_sha256s",
    "candidate_source_locations",
    "candidate_locator_grain",
    "candidate_snippet_or_cell_or_code_line",
    "candidate_parsed_value_review_only",
    "pass_rule_predicate",
    "locator_candidate_status",
    "pass_rule_adjudication_status",
    "parsed_candidate_admission_status",
    "exact_parse_blocker",
    "next_backend_action_after_parse",
]

FIELD_TERMS = {
    "source_cell_unit_sign__effective_contract_family_by_era": [
        "sofr",
        "eurodollar",
        "january 2022",
        "money market futures",
    ],
    "source_cell_unit_sign__literal_na_handling": ["literal_na", "not found", "na.omit"],
    "source_cell_unit_sign__percentage_point_basis_point_conversion": [
        "percentage points",
        "rate changes",
        "one-for-one",
    ],
    "source_cell_unit_sign__source_instrument_code": [
        "mp1",
        "mp2",
        "ff1",
        "ed1",
        "ed2",
        "ois1y",
        "ust2y",
    ],
    "source_cell_unit_sign__source_workbook_cell_unit": [
        "units",
        "percentage points",
        "mp1",
        "ed2",
    ],
    "bps_year_formula__horizon_weights": [
        "one-year",
        "ed2",
        "ed3",
        "ed4",
        "principal components",
        "pc1",
    ],
    "bps_year_formula__rate_change_unit_conversion": [
        "percentage points",
        "daily yield change",
        "rate changes",
    ],
    "bps_year_formula__sign_convention": ["pc1", "mps", "coef", "dy1", "normalize"],
    "event_date_horizon_grid__contract_reference_interval": [
        "sofr futures",
        "eurodollar futures",
        "ed1",
        "ed2",
        "date_time",
    ],
    "event_date_horizon_grid__event_date": ["date", "date_time", "fomc"],
    "event_date_horizon_grid__event_specific_horizon_start_end_dates": [
        "date_time",
        "sofr futures",
        "eurodollar futures",
    ],
    "event_date_horizon_grid__event_window": [
        "high-frequency",
        "around fomc",
        "30-minute",
    ],
    "event_date_horizon_grid__literal_na_exclusion": ["literal_na", "na.omit"],
    "loading_back_transform__factor_definition": [
        "principal component",
        "pc1",
        "prcomp",
    ],
    "loading_back_transform__instrument_loadings": [
        "mp1",
        "mp2",
        "ed2",
        "ed3",
        "ed4",
        "select(c(date",
    ],
    "loading_back_transform__rotation_sign_rule": ["pc1", "coef", "mps", "dy1"],
    "loading_back_transform__scalar_to_cell_back_transform": [
        "mps",
        "pc1",
        "coef",
        "normalized",
    ],
    "loading_back_transform__source_code_replication_command": [
        "source(\"mps.r\")",
        "run script",
        "mps.r",
    ],
}

PASS_RULES = {
    "source_cell_unit_sign": (
        "candidate must identify source-authored field unit/sign/instrument "
        "semantics at sheet-cell, code-line, or source-document-line grain"
    ),
    "bps_year_formula": (
        "candidate must identify a source-authored bps-year formula or all "
        "horizon weights and rate-change unit conversions at locator grain"
    ),
    "event_date_horizon_grid": (
        "candidate must identify event-date-specific horizon grid, event window, "
        "or explicit no-static-quarter rule at locator grain"
    ),
    "loading_back_transform": (
        "candidate must identify source-authored factor/loadings/back-transform "
        "or replication command at code-line or structured-cell grain"
    ),
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _clean(value: object, limit: int = 240) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", unescape(text)).strip()
    text = text.replace(";", ",")
    return text[:limit]


def _terms_for(field: str) -> list[str]:
    return [term.lower() for term in FIELD_TERMS.get(field, [])]


def _match(text: str, terms: list[str]) -> bool:
    lower = text.lower()
    for term in terms:
        if term == "literal_na":
            if re.search(r"(^|[,\s])NA($|[,\s])", text):
                return True
            continue
        if term in lower:
            return True
    return False


def _text_line_candidates(path: Path, field: str, *, member: str = "") -> list[dict[str, str]]:
    terms = _terms_for(field)
    if not terms:
        return []
    try:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    except UnicodeDecodeError:
        return []
    candidates: list[dict[str, str]] = []
    for line_no, line in enumerate(lines, start=1):
        if _match(line, terms):
            loc = f"{path.relative_to(ROOT).as_posix()}::line={line_no}"
            if member:
                loc = f"{path.relative_to(ROOT).as_posix()}::{member}::line={line_no}"
            candidates.append(
                {
                    "artifact": path.relative_to(ROOT).as_posix(),
                    "location": loc,
                    "grain": "source_document_line",
                    "snippet": _clean(line),
                    "value": "",
                }
            )
        if len(candidates) >= 8:
            break
    return candidates


def _html_candidates(path: Path, field: str) -> list[dict[str, str]]:
    terms = _terms_for(field)
    if not terms:
        return []
    text = path.read_text(encoding="utf-8", errors="replace")
    text = re.sub(r"<(script|style).*?</\1>", " ", text, flags=re.I | re.S)
    lines = [
        _clean(line)
        for line in re.sub(r"<[^>]+>", "\n", text).splitlines()
        if _clean(line)
    ]
    candidates: list[dict[str, str]] = []
    for line_no, line in enumerate(lines, start=1):
        if _match(line, terms):
            candidates.append(
                {
                    "artifact": path.relative_to(ROOT).as_posix(),
                    "location": f"{path.relative_to(ROOT).as_posix()}::text_line={line_no}",
                    "grain": "source_document_line",
                    "snippet": line,
                    "value": "",
                }
            )
        if len(candidates) >= 8:
            break
    return candidates


def _zip_candidates(path: Path, field: str) -> list[dict[str, str]]:
    terms = _terms_for(field)
    if not terms:
        return []
    candidates: list[dict[str, str]] = []
    with zipfile.ZipFile(path) as archive:
        for member in archive.namelist():
            if member.endswith("/"):
                continue
            try:
                text = archive.read(member).decode("utf-8", errors="replace")
            except Exception:
                continue
            for line_no, line in enumerate(text.splitlines(), start=1):
                if _match(line, terms):
                    candidates.append(
                        {
                            "artifact": path.relative_to(ROOT).as_posix(),
                            "location": (
                                f"{path.relative_to(ROOT).as_posix()}::{member}::line={line_no}"
                            ),
                            "grain": (
                                "code_line"
                                if member.lower().endswith((".r", ".py", ".m"))
                                else "source_document_line"
                            ),
                            "snippet": _clean(line),
                            "value": "",
                        }
                    )
                if len(candidates) >= 10:
                    return candidates
    return candidates


def _xlsx_candidates(path: Path, field: str) -> list[dict[str, str]]:
    terms = _terms_for(field)
    if not terms:
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    candidates: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        max_row = min(sheet.max_row or 0, 12)
        max_col = min(sheet.max_column or 0, 40)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
        ):
            for cell in row:
                value = cell.value
                text = _clean(value)
                if text and _match(text, terms):
                    candidates.append(
                        {
                            "artifact": path.relative_to(ROOT).as_posix(),
                            "location": (
                                f"{path.relative_to(ROOT).as_posix()}::{sheet.title}!{cell.coordinate}"
                            ),
                            "grain": "sheet_cell",
                            "snippet": text,
                            "value": text,
                        }
                    )
                if len(candidates) >= 10:
                    return candidates
    return candidates


def _pdf_candidates(path: Path, field: str) -> list[dict[str, str]]:
    terms = _terms_for(field)
    if not terms:
        return []
    candidates: list[dict[str, str]] = []
    try:
        reader = PdfReader(str(path))
    except Exception:
        return []
    for page_no, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            continue
        for line_no, line in enumerate(text.splitlines(), start=1):
            if _match(line, terms):
                candidates.append(
                    {
                        "artifact": path.relative_to(ROOT).as_posix(),
                        "location": (
                            f"{path.relative_to(ROOT).as_posix()}::pdf_page={page_no}::text_line={line_no}"
                        ),
                        "grain": "source_document_line",
                        "snippet": _clean(line),
                        "value": "",
                    }
                )
            if len(candidates) >= 6:
                return candidates
    return candidates


def _artifact_candidates(path_text: str, field: str) -> list[dict[str, str]]:
    path = ROOT / path_text
    if not path.exists():
        return []
    suffix = path.suffix.lower()
    if suffix in {".html", ".htm"}:
        return _html_candidates(path, field)
    if suffix == ".zip":
        return _zip_candidates(path, field)
    if suffix == ".xlsx":
        return _xlsx_candidates(path, field)
    if suffix == ".pdf":
        return _pdf_candidates(path, field)
    if suffix in {".csv", ".md", ".txt", ".r"}:
        return _text_line_candidates(path, field)
    return []


def _join(values: list[str]) -> str:
    return ";".join(dict.fromkeys(value for value in values if value))


def main() -> int:
    if not INPUT.exists():
        print(f"missing input: {INPUT}", file=sys.stderr)
        return 1
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with INPUT.open(encoding="utf-8", newline="") as handle:
        packet_rows = list(csv.DictReader(handle))

    rows: list[dict[str, str]] = []
    for packet in packet_rows:
        field = packet.get("authored_field_name", "")
        artifacts = [
            path
            for path in packet.get("downloaded_artifact_paths", "").split(";")
            if path
        ]
        candidates: list[dict[str, str]] = []
        for artifact in artifacts:
            for candidate in _artifact_candidates(artifact, field):
                candidates.append(candidate)
                if len(candidates) >= 12:
                    break
            if len(candidates) >= 12:
                break

        if candidates:
            parse_class = "downloaded_public_artifact_locator_candidate_parse"
            locator_status = "pass_locator_grain_candidate_extracted_review_only"
            adjudication_status = (
                "blocked_pass_rule_adjudication_pending_sibling_gates_not_admission"
            )
            parsed_status = "blocked_parsed_locator_candidate_review_only_not_field_evidence"
            blocker = (
                f"{field} has downloaded-artifact locator candidates, but "
                "parsed snippets/cells/code lines are review-only until the "
                "field pass rule, authored-invariant sibling gates, and "
                "independent-replication sibling gates all pass."
            )
            next_action = (
                "adjudicate_locator_candidate_against_field_pass_rule_then_resolve_sibling_gates_fail_closed"
            )
        elif artifacts:
            parse_class = "downloaded_public_artifact_locator_candidate_no_hit"
            locator_status = "blocked_no_locator_grain_candidate_extracted"
            adjudication_status = "blocked_no_candidate_available_for_pass_rule"
            parsed_status = "blocked_no_parsed_candidate_admitted"
            blocker = (
                f"{field} has downloaded artifacts, but the bounded parser did "
                "not find a locator-grain candidate satisfying the review terms."
            )
            next_action = "manual_review_downloaded_artifacts_or_acquire_new_source_family"
        else:
            parse_class = "manual_authenticated_new_source_family_parse_blocker"
            locator_status = "blocked_manual_authenticated_or_new_source_family_required"
            adjudication_status = "blocked_no_downloaded_artifact_for_pass_rule"
            parsed_status = "blocked_no_parsed_candidate_admitted"
            blocker = (
                f"{field} remains a manual-authenticated/new-source-family "
                "blocker; no downloaded public artifact exists to parse."
            )
            next_action = "manual_authenticated_or_new_source_family_acquisition_required"

        protocol_component = packet.get("protocol_component", "")
        rows.append(
            {
                "policy_path_downloaded_artifact_locator_parse_adjudication_manifest_row_id": (
                    f"policy_path_downloaded_artifact_locator_parse_adjudication_manifest::{len(rows) + 1:04d}"
                ),
                "policy_path_real_source_author_web_acquisition_attempt_packet_row_id": packet.get(
                    "policy_path_real_source_author_web_acquisition_attempt_packet_row_id",
                    "",
                ),
                "authored_field_name": field,
                "parse_attempt_class": parse_class,
                "candidate_locator_count": str(len(candidates)),
                "candidate_source_artifact_paths": _join(
                    [candidate["artifact"] for candidate in candidates]
                ),
                "candidate_source_artifact_sha256s": _join(
                    [_sha256(ROOT / candidate["artifact"]) for candidate in candidates]
                ),
                "candidate_source_locations": _join(
                    [candidate["location"] for candidate in candidates]
                ),
                "candidate_locator_grain": _join(
                    [candidate["grain"] for candidate in candidates]
                ),
                "candidate_snippet_or_cell_or_code_line": _join(
                    [candidate["snippet"] for candidate in candidates]
                ),
                "candidate_parsed_value_review_only": _join(
                    [candidate["value"] for candidate in candidates]
                ),
                "pass_rule_predicate": PASS_RULES.get(
                    protocol_component,
                    "candidate must satisfy source-authored locator-grain field pass rule",
                ),
                "locator_candidate_status": locator_status,
                "pass_rule_adjudication_status": adjudication_status,
                "parsed_candidate_admission_status": parsed_status,
                "exact_parse_blocker": blocker,
                "next_backend_action_after_parse": next_action,
            }
        )

    with MANIFEST.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    print(MANIFEST.relative_to(ROOT).as_posix())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
